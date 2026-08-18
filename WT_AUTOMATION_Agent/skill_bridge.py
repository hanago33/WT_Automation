# encoding: utf-8
"""WT_AUTOMATION_Agent Skill 集成桥 —— 加载和解析 Qoder Skill 定义。

Skill 是 Qoder 生态中用于扩展 Agent 能力的知识包。
本模块负责从文件系统加载 Skill 内容，转换为 Agent 可用的上下文。
"""
from __future__ import annotations

import json
import os
import re
from typing import Any


# ---------------------------------------------------------------------------
# Skill 结构定义
# ---------------------------------------------------------------------------

SKILL_METADATA_FIELDS = {
    "name": "名称",
    "description": "描述",
    "version": "版本",
    "author": "作者",
}


class SkillInfo:
    """Skill 信息。"""
    def __init__(
        self,
        name: str = "",
        description: str = "",
        content: str = "",
        file_path: str = "",
        metadata: dict[str, str] | None = None,
    ):
        self.name = name
        self.description = description
        self.content = content
        self.file_path = file_path
        self.metadata = metadata or {}

    def to_context_text(self) -> str:
        """转换为 Agent 可用的上下文文本。"""
        parts = [f"Skill: {self.name}"]
        if self.description:
            parts.append(f"描述: {self.description}")
        if self.content:
            parts.append("---")
            parts.append(self.content)
        return "\n".join(parts)


# ---------------------------------------------------------------------------
# 从文件系统加载 Skill
# ---------------------------------------------------------------------------

def discover_skills(skill_dirs: list[str]) -> list[SkillInfo]:
    """从多个目录中发现 Skill 文件。

    支持格式：
    - *.md 文件（SKILL.md / 任意 .md）
    - *.json 文件（结构化 skill 定义）
    - *.xlsx / *.csv 文件（表格型 Skill，自动转为可读文本）
    """
    skills: list[SkillInfo] = []
    seen: set[str] = set()

    for skill_dir in skill_dirs:
        if not os.path.isdir(skill_dir):
            continue
        for root, _dirs, files in os.walk(skill_dir):
            for file_name in files:
                ext = os.path.splitext(file_name)[1].lower()
                if ext not in (".md", ".json", ".xlsx", ".xls", ".csv", ".tsv", ".pdf", ".docx"):
                    continue
                file_path = os.path.join(root, file_name)
                if file_path in seen:
                    continue
                seen.add(file_path)

                try:
                    skill = _load_skill_file(file_path)
                    if skill:
                        skills.append(skill)
                except Exception as exc:
                    import logging
                    logging.debug("加载 Skill 文件失败 %s: %s", file_path, exc)

    return skills


def _load_skill_file(file_path: str) -> SkillInfo | None:
    """加载单个 Skill 文件。"""
    ext = os.path.splitext(file_path)[1].lower()
    name = os.path.splitext(os.path.basename(file_path))[0]

    if ext == ".json":
        return _load_json_skill(file_path, name)
    elif ext == ".md":
        return _load_md_skill(file_path, name)
    elif ext in (".xlsx", ".xls"):
        return _load_xlsx_skill(file_path, name)
    elif ext in (".csv", ".tsv", ".tab"):
        return _load_csv_skill(file_path, name)
    elif ext == ".pdf":
        return _load_pdf_skill(file_path, name)
    elif ext == ".docx":
        return _load_docx_skill(file_path, name)
    return None


def _load_json_skill(file_path: str, default_name: str) -> SkillInfo | None:
    """加载 JSON 格式的 Skill。"""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError):
        return None

    if isinstance(data, dict):
        name = str(data.get("name", data.get("title", default_name)))
        description = str(data.get("description", data.get("desc", "")))
        content = json.dumps(data, ensure_ascii=False, indent=2)
        metadata = {k: str(data.get(k, "")) for k in SKILL_METADATA_FIELDS if k in data}
        return SkillInfo(
            name=name,
            description=description,
            content=content,
            file_path=file_path,
            metadata=metadata,
        )
    return None


def _load_md_skill(file_path: str, default_name: str) -> SkillInfo | None:
    """加载 Markdown 格式的 Skill。"""
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
    except OSError:
        return None

    name = default_name
    description = ""

    # 尝试从 front-matter 或首行提取元数据
    lines = content.splitlines()
    for line in lines[:10]:
        line_stripped = line.strip()
        if line_stripped.startswith("# "):
            name = line_stripped.lstrip("# ").strip()
        elif line_stripped.startswith("> "):
            description = line_stripped.lstrip("> ").strip()
        # 支持 YAML front-matter 格式
        m = re.match(r"^name:\s*(.+)$", line_stripped, re.IGNORECASE)
        if m:
            name = m.group(1).strip()
        m = re.match(r"^description:\s*(.+)$", line_stripped, re.IGNORECASE)
        if m:
            description = m.group(1).strip()

    return SkillInfo(
        name=name,
        description=description,
        content=content,
        file_path=file_path,
    )


def _load_xlsx_skill(file_path: str, default_name: str) -> SkillInfo | None:
    """加载 .xlsx/.xls 格式 Skill（表格型知识）。"""
    try:
        import openpyxl
    except ImportError:
        return SkillInfo(
            name=default_name,
            description="Excel 知识表（需 openpyxl 库读取，pip install openpyxl）",
            content=f"文件: {file_path}\n\n[需要 pip install openpyxl 才能读取此 Skill]",
            file_path=file_path,
        )
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception:
        return None

    name = default_name
    description = "Excel 知识表"
    lines: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"## Sheet: {sheet_name}")
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            if row_count >= 50:  # 每 sheet 最多 50 行
                lines.append(f"... (共 {ws.max_row} 行，已截断前 50 行)")
                break
            cells = [str(c).strip() if c is not None else "" for c in row]
            lines.append(" | ".join(cells))
            row_count += 1
        lines.append("")
    wb.close()

    return SkillInfo(
        name=name,
        description=description,
        content="\n".join(lines),
        file_path=file_path,
    )


def _load_csv_skill(file_path: str, default_name: str) -> SkillInfo | None:
    """加载 .csv/.tsv 格式 Skill。"""
    import csv
    try:
        with open(file_path, "r", encoding="utf-8-sig") as f:
            delimiter = "\t" if file_path.lower().endswith((".tsv", ".tab")) else ","
            reader = csv.reader(f, delimiter=delimiter)
            lines: list[str] = []
            for ri, row in enumerate(reader):
                if ri >= 50:
                    lines.append(f"... (已截断前 50 行)")
                    break
                cells = [c.strip() for c in row]
                lines.append(" | ".join(cells))
    except Exception:
        return None

    return SkillInfo(
        name=default_name,
        description="CSV 知识表",
        content="\n".join(lines),
        file_path=file_path,
    )


def _load_pdf_skill(file_path: str, default_name: str) -> SkillInfo | None:
    """加载 PDF 知识文档（返回占位 Skill，实际解析留给后续 AI Agent）。"""
    return SkillInfo(
        name=default_name,
        description="PDF 领域手册（由 Help_document 提供，运行时按需解析）",
        content=(
            f"文件: {os.path.basename(file_path)}\n"
            f"[此 PDF 的知识内容将在 Agent 调用时按需提取，"
            f"当前作为可发现的领域知识源注册]"
        ),
        file_path=file_path,
    )


def _load_docx_skill(file_path: str, default_name: str) -> SkillInfo | None:
    """加载 DOCX 知识文档（返回占位 Skill，实际解析留给后续 AI Agent）。"""
    return SkillInfo(
        name=default_name,
        description="DOCX 项目文档（由 Help_document 提供，运行时按需解析）",
        content=(
            f"文件: {os.path.basename(file_path)}\n"
            f"[此 DOCX 的知识内容将在 Agent 调用时按需提取]"
        ),
        file_path=file_path,
    )


# ---------------------------------------------------------------------------
# 内置领域 Skill（免加载即可使用）
# ---------------------------------------------------------------------------

WIN_UI_AUTOMATION_SKILL_CONTENT = """
## Windows UI 自动化最佳实践
1. 优先使用 automation_id 定位控件，比 name/xpath 更稳定。
2. WPF 应用使用 UIA (UI Automation) 框架，Win32 应用使用 Win32 API。
3. 对于自绘控件，使用相对区域点击 (click_relative_region) 替代控件定位。
4. 控件查找建议顺序：automation_id > name+control_type > class_name > template_match。
5. 等待控件出现后再操作，超时建议 3-5 秒。
6. 操作完成后适当等待界面响应（waitAfter 0.3-0.5 秒）。
7. 失败重试 1-2 次，间隔 1 秒。
"""


WT_AUTOMATION_LESSONS_SKILL_CONTENT = """
## WT 自动化调试根因与工程改进（项目实战沉淀）
心智模型：
- "动作成功" ≠ "业务生效"：日志"点击/输入成功"只代表控件被操作，必须校验界面状态变化（窗口出现/值提交/行新增）。多数疑难 bug 是"假成功"。
- 报错步常是连带受害者：先回溯到"第一个界面没按预期变化"的步骤修复，别在报错步打补丁。
- 运行时定位是"全遍历+打分取最高"，非条件查询：评分过宽会误命中，缓存未校验归属会命中错对象。
- 定位分层降级：复合精确→ui_path 尾部对齐→模板图像→坐标；命中靠后层视为退化告警。

运行时根因模式（症状→修复）：
- 假成功·命中文本展示层（Text/TextBlock、IsKeyboardFocusable=False、无 Invoke/Value/Toggle）→ 换真实可交互目标或 click_relative_region，降权展示型。
- 假成功·输入未提交 → 输入动作补 postInputKeys:{TAB} 或 {ENTER} 强制提交失焦。
- 错窗口·无标题主窗冒充弹窗（落到 Window_Main 0,0,大 rect）→ 前台窗口优先，空标题仅前台兜底。
- 矩形基准漂移（整段相对区域统一偏移）→ 优先原生 GetWindowRect 取稳定基准。
- 缓存污染 → 缓存命中处加窗口归属校验。
- 同名同类误命中 → 保留录制序号 foundIndex，运行时 get_wrapper_found_index 作"最低优先回退候选 + +12 消歧"，绝不覆盖可靠 id/name。

工程规则：
- 复合定位器 method/value 逗号成对、运行时 AND 匹配；优先级 automation_id,control_type > automation_id > ui_path(深度≥2) > name,control_type > name > class_name > control_type,found_index。
- 动态控件用组合(id+type/name+type)稳住，深嵌套用 ui_path 尾部对齐，最后才 found_index/模板/坐标。
- fallbackChain 四级：L1主定位→L2 ui_path_search→L3 template→L4 coordinate；非 L1 命中=退化，须记反馈+自愈。

字段契约纪律（易踩静默坑）：
- normalize_step 白名单重建时，新增字段（fallbackChain、_ 前缀内部字段）必须显式放行，否则被静默丢弃。
- Excel 往返改字段=列定义/规范化/写/读/清理五处同步；公共入口签名改动要同步所有调用点。
完整人读版：docs/典型问题记录/WT自动化_调试根因与工程改进知识库.md
"""


WT_CONTROL_SEMANTICS_SKILL_CONTENT = """
## WT 折叠面板与参数控件语义速查（Agent 生成步骤必读）

1. **"XX参数"多为折叠面板切换按钮，不是输入控件。**
   WT 中"求解器参数""风电场参数""建模区域参数"等名称通常是可折叠面板的切换按钮
   （automationId 形如 MTDTileView_Button_ToggleState / ExpanderWithToolBar_Button_ToggleState，
   controlType=Button），点击后才展开对应的 View（如 MUPDASSolverParametersView），
   面板**内部**才是真正的参数输入框/下拉框。
   → 用户说"把 XX 参数设置为 Y"时，应输出两个步骤：
     ① click 折叠面板切换按钮（展开面板）；
     ② 在面板内对具体参数控件做 type_text / set_combobox / click。

2. **求解器参数面板（MUPDASSolverParametersView，位于 CFD 新建流程 MUPDASCreatorView 内）映射：**
   - 展开按钮：`求解器参数` → targetValue 形如 `MTDTileView_Button_ToggleState,Button`
   - 最大迭代次数（默认 25）→ `DASParameters_NumericUpDownWithWarning_NbIterations,Custom`（Edit）
   - 收敛阈值（默认 0.98 等）→ `DASParameters_NumericUpDownWithWarning_ConvergenceThreshold,Custom`（Edit）
   - 并行线程数 → `DASParameters_NumericUpDownWithWarning_NumberThreads,...`（Edit）
   - 用户说"求解器参数设置为 4"时，默认按"最大迭代次数=4"解释；
     若上下文无提示具体参数，可先给出该默认解释并在描述中注明。

3. **NumericUpDown（className MUPNumericUpDownWithWarning）输入后必须提交失焦：**
   type_text 之后追加一个 send_keys 步骤（text 填 {TAB} 或 {ENTER}），
   或用 type_text_relative 的 postInputKeys，否则值可能未写入。

4. **控件检索选型规则：**
   交互动作（click / type_text / set_combobox / select_dropdown_item_runtime）必须选中
   Button / Edit / ComboBox / ListBoxItem 等可交互类型；**禁止选 Text / TextBlock 文字展示层**
   （名称完全相同的文字层往往排在控件子节点里，点击会"假成功"）。
   同名控件（同一 automationId 多实例）优先选 labelText / uiPath 与当前视图匹配的候选，
   必要时用 find_control 的 within 参数限定窗口/视图。
"""


def get_builtin_skills() -> list[SkillInfo]:
    """获取内置 Skill 列表（无需加载文件）。"""
    return [
        SkillInfo(
            name="Windows UI Automation",
            description="Windows 桌面 UI 自动化最佳实践指南",
            content=WIN_UI_AUTOMATION_SKILL_CONTENT,
        ),
        SkillInfo(
            name="WT Automation Lessons",
            description="WT 自动化调试根因与工程改进经验（假成功/定位/降级/字段契约）",
            content=WT_AUTOMATION_LESSONS_SKILL_CONTENT,
        ),
        SkillInfo(
            name="WT Control Semantics",
            description="WT 折叠面板与参数控件语义速查（求解器参数面板映射、NumericUpDown 提交、控件选型）",
            content=WT_CONTROL_SEMANTICS_SKILL_CONTENT,
        ),
    ]


# ---------------------------------------------------------------------------
# 便捷函数
# ---------------------------------------------------------------------------

def load_all_skills_text(
    additional_dirs: list[str] | None = None,
    include_builtin: bool = True,
) -> str:
    """加载所有可用 Skill 并合并为一段文本。

    参数：
        additional_dirs: 额外的 Skill 文件目录
        include_builtin: 是否包含内置 Skill

    返回：
        合并后的纯文本
    """
    all_skills: list[SkillInfo] = []
    if include_builtin:
        all_skills.extend(get_builtin_skills())

    if additional_dirs:
        all_skills.extend(discover_skills(additional_dirs))

    if not all_skills:
        return ""

    parts: list[str] = []
    for skill in all_skills:
        text = skill.to_context_text()
        if text:
            parts.append(text)
    return "\n\n---\n\n".join(parts)
