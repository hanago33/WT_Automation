# -*- coding: utf-8 -*-
"""
生成「10min 时序风速 -> 逐时功率」Excel 计算模板。

物理模型（空气密度修正，IEC 常用等效风速法）：
  1. 将实际空气密度下的风速，折算到标准空气密度(1.225 kg/m3)下的等效风速：
        v_std = v_actual * (rho / rho0) ** (1/3)
  2. 用标准功率曲线(风速->功率)对 v_std 线性插值，得到该 10min 的功率。
  3. 逐时：对该小时内 6 个 10min 功率取平均，得到逐时功率(kW)与逐时发电量(kWh)。

依赖: openpyxl
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RHO0 = 1.225  # 标准空气密度 kg/m3

OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                         "逐时功率计算模板.xlsx")

# ---------- 样式 ----------
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")   # 黄色=输入
CALC_FILL = PatternFill("solid", fgColor="E2EFDA")    # 绿色=计算
PARAM_FILL = PatternFill("solid", fgColor="DDEBF7")   # 蓝色=参数
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
SUB_FONT = Font(italic=True, size=9, color="595959")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def box(ws, r1, c1, r2, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = BORDER


# ============================================================
# 工作簿
# ============================================================
wb = openpyxl.Workbook()

# ---------- 页1: 说明 ----------
ws_info = wb.active
ws_info.title = "说明"
ws_info.column_dimensions["A"].width = 100
info_lines = [
    ("逐时功率计算模板", TITLE_FONT),
    ("", None),
    ("【用途】输入 10 分钟时序风速、标准功率曲线、空气密度，自动计算逐时功率与发电量。", None),
    ("", None),
    ("【黄色单元格 = 需要你填写的输入】", Font(bold=True, color="BF8F00")),
    ("   • 标准空气密度(kg/m3): 默认 1.225", None),
    ("   • 空气密度: 可在[输入数据]页填单一值，或按 10min 逐行填写（优先用逐行值）", None),
    ("   • 10min 风速(m/s): 一天 144 行（24h × 6）", None),
    ("   • 标准功率曲线: 风速(m/s) -> 功率(kW)，按升序，至少包含切入/额定/切出点", None),
    ("", None),
    ("【绿色单元格 = 自动计算，请勿手改】", Font(bold=True, color="375623")),
    ("", None),
    ("【计算方法 - 空气密度等效风速修正(IEC常用)】", Font(bold=True)),
    ("   1. 等效标准密度风速:  v_std = v_actual × (ρ / ρ0)^(1/3)", None),
    ("   2. 用标准功率曲线对 v_std 线性插值得到该 10min 功率(kW)", None),
    ("   3. 逐时功率 = 该小时内 6 个 10min 功率的平均值(kW)", None),
    ("   4. 逐时发电量 = 逐时功率 × 1(h) = kWh", None),
    ("   注: 风速低于切入或高于切出时功率为 0; 高于额定风速时取额定功率(曲线最后非零段)。", None),
    ("", None),
    ("【提示】功率曲线插值使用 TREND/FORECAST 不便处理分段，模板用 LOOKUP+INDEX 查找方式在[输入数据]页实现。", None),
]
for i, (txt, font) in enumerate(info_lines, start=1):
    cell = ws_info.cell(row=i, column=1, value=txt)
    if font:
        cell.font = font
    cell.alignment = WRAP

# ---------- 页2: 输入数据 ----------
ws = wb.create_sheet("输入数据")

# 参数区(顶部)
ws.cell(row=1, column=1, value="参数").font = TITLE_FONT
ws.cell(row=2, column=1, value="标准空气密度 ρ0 (kg/m3)").font = Font(bold=True)
ws.cell(row=2, column=2, value=RHO0).fill = PARAM_FILL
ws.cell(row=2, column=2).border = BORDER
ws.cell(row=3, column=1, value="默认空气密度 ρ (kg/m3)").font = Font(bold=True)
ws.cell(row=3, column=2, value=1.225).fill = INPUT_FILL   # 单行默认值
ws.cell(row=3, column=2).border = BORDER
ws.cell(row=3, column=3,
        value="若不逐行填写空气密度，则全部使用此值").font = SUB_FONT

PARAM_RHO0 = "$B$2"
PARAM_RHO_DEF = "$B$3"

# 功率曲线表 (放在 H/I 列，避免与时序区 A-G 列冲突)
ws.cell(row=1, column=8, value="标准功率曲线 (风速 m/s -> 功率 kW)").font = TITLE_FONT
ws.cell(row=2, column=8, value="风速(m/s)").font = HDR_FONT
ws.cell(row=2, column=8).fill = HDR_FILL
ws.cell(row=2, column=9, value="功率(kW)").font = HDR_FONT
ws.cell(row=2, column=9).fill = HDR_FILL
ws.cell(row=2, column=8).alignment = CENTER
ws.cell(row=2, column=9).alignment = CENTER
ws.cell(row=2, column=8).border = BORDER
ws.cell(row=2, column=9).border = BORDER

# 示例标准功率曲线(某 2MW 机型)，用户可替换
sample_curve = [
    (0, 0), (3, 0), (3.5, 60), (4, 130), (5, 280), (6, 470),
    (7, 700), (8, 980), (9, 1250), (10, 1520), (11, 1720),
    (12, 1880), (13, 1960), (14, 2000), (15, 2000), (25, 2000), (26, 0),
]
for i, (v, p) in enumerate(sample_curve, start=3):
    ws.cell(row=i, column=8, value=v).fill = INPUT_FILL
    ws.cell(row=i, column=9, value=p).fill = INPUT_FILL
    ws.cell(row=i, column=8).border = BORDER
    ws.cell(row=i, column=9).border = BORDER
CURVE_FIRST = 3
CURVE_LAST = CURVE_FIRST + len(sample_curve) - 1
CURVE_V = f"$H${CURVE_FIRST}:$H${CURVE_LAST}"
CURVE_P = f"$I${CURVE_FIRST}:$I${CURVE_LAST}"

ws.column_dimensions["H"].width = 12
ws.column_dimensions["I"].width = 12

# 时序数据表 (左侧, 从第6行开始)
ws.cell(row=5, column=1, value="10 分钟时序数据（一天 144 段）").font = TITLE_FONT
headers = ["序号", "小时", "10min序号", "风速(m/s)", "空气密度(kg/m3,可选)",
           "等效标准风速(m/s)", "10min功率(kW)"]
for j, h in enumerate(headers, start=1):
    ws.cell(row=6, column=j, value=h)
style_header(ws, 6, len(headers))

DATA_FIRST = 7
DATA_LAST = DATA_FIRST + 143  # 144 行

for idx, r in enumerate(range(DATA_FIRST, DATA_LAST + 1)):
    n = idx + 1
    hour = (n - 1) // 6       # 0..23
    sub = (n - 1) % 6 + 1     # 1..6
    ws.cell(row=r, column=1, value=n)
    ws.cell(row=r, column=2, value=hour)
    ws.cell(row=r, column=3, value=sub)
    # 风速 输入
    c_v = ws.cell(row=r, column=4, value=None)
    c_v.fill = INPUT_FILL
    # 空气密度(可选) 输入
    c_rho = ws.cell(row=r, column=5, value=None)
    c_rho.fill = INPUT_FILL
    # 等效标准风速 = v * (rho/rho0)^(1/3)
    ws.cell(row=r, column=6,
            value=(f'=IF(D{r}="","",D{r}*'
                   f'IF(E{r}="",{PARAM_RHO_DEF},{PARAM_RHO_DEF})'
                   f'*IF(E{r}<>"",E{r}/{PARAM_RHO0},1))'
                   # 修正：若 E 有值用 E，否则用默认
                   ))
    # 上面的等效风速公式用更稳妥写法(下面重写)
box(ws, DATA_FIRST, 1, DATA_LAST, len(headers))

# 重写 等效风速 列(F) 公式：优先用 E 行密度，否则用默认密度
for r in range(DATA_FIRST, DATA_LAST + 1):
    ws.cell(row=r, column=6,
            value=(f'=IF(D{r}="","",D{r}*'
                   f'IF(E{r}<>"",E{r},{PARAM_RHO_DEF})/{PARAM_RHO0}'
                   f')^(1/3))'))

# 10min 功率列(G)：对等效风速在功率曲线上线性插值，含切入/切出/额定处理
# 思路:
#  v_std < 曲线首个风速 -> 0
#  v_std > 曲线末个风速 -> 0 (切出)
#  介于之间 -> TREND 插值(分段线性可由 TREND 在两点间实现, 这里用 FORECAST 近似:
#     取相邻两点做线性插值。用 LOOKUP 找下界，再算斜率)
# 为稳健，采用: 用 LOOKUP 取 <=v_std 的最大风速点及其功率，再取下一个点做线性插值。
# 功率曲线点数
N_CURVE = CURVE_LAST - CURVE_FIRST + 1
for r in range(DATA_FIRST, DATA_LAST + 1):
    f_vstd = f"F{r}"
    # 下界风速 / 功率（LOOKUP 找 <=v_std 的最大点）
    below_v = f'LOOKUP({f_vstd},{CURVE_V})'
    below_p = f'LOOKUP({f_vstd},{CURVE_V},{CURVE_P})'
    # 上界：下界点之后一点
    above_idx = f'(MATCH({f_vstd},{CURVE_V},1)+1)'
    above_v = f'INDEX({CURVE_V},{above_idx})'
    above_p = f'INDEX({CURVE_P},{above_idx})'
    # 逻辑：空->""; <首个风速或>末风速 -> 0(切出); 否则线性插值
    formula = (
        f'=IF({f_vstd}="","",'
        f'IF(OR({f_vstd}<=INDEX({CURVE_V},1),{f_vstd}>=INDEX({CURVE_V},{N_CURVE})),0,'
        f'IF({above_idx}>{N_CURVE},0,'
        f'{below_p}+({above_p}-{below_p})*({f_vstd}-{below_v})/({above_v}-{below_v})'
        f')))'
    )
    ws.cell(row=r, column=7, value=formula).fill = CALC_FILL

ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 6
ws.column_dimensions["C"].width = 8
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 16
ws.column_dimensions["F"].width = 16
ws.column_dimensions["G"].width = 14

# ---------- 页3: 逐时功率 ----------
ws_h = wb.create_sheet("逐时功率")
ws_h.cell(row=1, column=1, value="逐时功率与发电量（按小时聚合 6 个 10min）").font = TITLE_FONT
h_headers = ["小时", "平均风速(m/s)", "逐时功率(kW)", "逐时发电量(kWh)", "备注"]
for j, h in enumerate(h_headers, start=1):
    ws_h.cell(row=2, column=j, value=h)
style_header(ws_h, 2, len(h_headers))

HOUR_FIRST = 3
HOUR_LAST = HOUR_FIRST + 23
for h in range(24):
    r = HOUR_FIRST + h
    ws_h.cell(row=r, column=1, value=h)
    # 该小时对应 输入数据 的行范围
    d_start = DATA_FIRST + h * 6
    d_end = d_start + 5
    # 平均风速(仅对非空风速平均)
    ws_h.cell(row=r, column=2,
              value=(f'=IFERROR(AVERAGE(输入数据!D{d_start}:D{d_end}),"")'))
    # 逐时功率 = 6个10min功率平均
    ws_h.cell(row=r, column=3,
              value=(f'=IFERROR(AVERAGE(输入数据!G{d_start}:G{d_end}),"")')
              ).fill = CALC_FILL
    # 逐时发电量 = 逐时功率 * 1h
    ws_h.cell(row=r, column=4,
              value=(f'=IF(C{r}="","",C{r}*1)')).fill = CALC_FILL
    ws_h.cell(row=r, column=5, value="")

# 合计行
tot = HOUR_LAST + 1
ws_h.cell(row=tot, column=1, value="全天合计").font = Font(bold=True)
ws_h.cell(row=tot, column=4,
          value=f'=SUM(D{HOUR_FIRST}:D{HOUR_LAST})').font = Font(bold=True)
ws_h.cell(row=tot, column=4).fill = CALC_FILL

box(ws_h, HOUR_FIRST, 1, HOUR_LAST, len(h_headers))
box(ws_h, tot, 1, tot, 4)

ws_h.column_dimensions["A"].width = 8
ws_h.column_dimensions["B"].width = 14
ws_h.column_dimensions["C"].width = 14
ws_h.column_dimensions["D"].width = 16
ws_h.column_dimensions["E"].width = 10

# 冻结窗格
ws.freeze_panes = "A7"
ws_h.freeze_panes = "A3"

wb.save(OUT_PATH)
print("已生成模板:", OUT_PATH)
