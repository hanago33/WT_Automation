# encoding: utf-8
"""
text_tools.py —— 文本处理工具（双卡片切换）

卡片一：TXT 合并
  - 选择多个 txt 文件 / 文件夹，按列表顺序合并成一个 txt 文档
  - 合并方式：直接接在末尾 / 新起一行再接
  - 可选：文件名作为章节标题、自定义分隔符、去重空行、按文件名排序
  - 输出编码可选：UTF-8 / GBK
  - 支持拖拽添加文件（需 tkinterdnd2，未安装时自动降级）

卡片二：CSV 转换
  - 选择多个 CSV 文件 / 文件夹
  - 自动识别编码（UTF-16/UTF-8/GBK 等）与分隔符（逗号/分号/制表符）
  - 转换类型：转 XLSX / 转 TXT
  - 转 XLSX 可选合并为单个工作簿（多工作表）
  - 进度条 + 日志

用法：
  1. 双击运行，或命令行执行：python text_tools.py
  2. 在顶部切换卡片
  3. 按卡片内提示操作
"""

import os
import sys
import csv
import io
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# 常见文本编码，按优先级尝试解码
_ENCODINGS = ["utf-8", "gbk", "gb18030", "utf-16", "big5", "latin-1"]

# 拖拽支持：优先用 tkinterdnd2，否则用 pywin32 的 OLE 拖拽
_HAS_DND = False
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    _HAS_DND = True
except Exception:  # noqa: BLE001
    pass

_HAS_OLE_DND = False
try:
    import pythoncom
    from win32com.shell import shell, shellcon
    _HAS_OLE_DND = True
except Exception:  # noqa: BLE001
    pass


# =====================================================================
# 卡片一：TXT 合并
# =====================================================================

class OleDropTarget:
    """基于 pywin32 的 OLE 文件拖拽目标（IDropTarget 实现，备用方案）。"""

    _com_interfaces_ = [pythoncom.IID_IDropTarget]
    _public_methods_ = ["DragEnter", "DragOver", "DragLeave", "Drop"]

    def __init__(self, hwnd, on_files):
        self._hwnd = hwnd
        self._on_files = on_files
        self._registered = False

    def register(self):
        try:
            pythoncom.RegisterDragDrop(self._hwnd, self)
            self._registered = True
            return True
        except Exception:  # noqa: BLE001
            return False

    def unregister(self):
        try:
            if self._registered:
                pythoncom.RevokeDragDrop(self._hwnd)
                self._registered = False
        except Exception:  # noqa: BLE001
            pass

    def DragEnter(self, pDataObj, grfKeyState, pt, pdwEffect):
        pdwEffect[0] = shellcon.DROPEFFECT_COPY
        return 0

    def DragOver(self, grfKeyState, pt, pdwEffect):
        pdwEffect[0] = shellcon.DROPEFFECT_COPY
        return 0

    def DragLeave(self):
        return 0

    def Drop(self, pDataObj, grfKeyState, pt, pdwEffect):
        pdwEffect[0] = shellcon.DROPEFFECT_COPY
        try:
            files = []
            try:
                data = pDataObj.GetData(shellcon.CF_HDROP)
                files = shell.DragQueryFile(data)
            except Exception:  # noqa: BLE001
                pass
            if files:
                self._on_files(files)
        except Exception:  # noqa: BLE001
            pass
        return 0


def read_text_file(path):
    """读取文本文件内容，自动探测编码，返回 (内容, 编码)。"""
    with open(path, "rb") as f:
        raw = f.read()
    for enc in ["utf-8-sig"] + _ENCODINGS:
        try:
            return raw.decode(enc), enc
        except (UnicodeDecodeError, LookupError):
            continue
    return raw.decode("latin-1"), "latin-1"


def merge_txt_files(
    file_paths,
    output_path,
    newline_between=False,
    add_headers=False,
    separator="",
    remove_empty_lines=False,
    output_encoding="utf-8-sig",
):
    """按顺序合并多个 txt 文件到输出文件，保持内容原样。"""
    parts = []
    for i, path in enumerate(file_paths):
        content, _ = read_text_file(path)
        if i > 0:
            if separator:
                parts.append(separator + "\n")
            elif newline_between:
                parts.append("\n")
        if add_headers:
            parts.append(os.path.basename(path) + "\n")
        parts.append(content)
    merged = "".join(parts)
    if remove_empty_lines:
        merged = "\n".join(line for line in merged.split("\n") if line.strip() != "")
    with open(output_path, "w", encoding=output_encoding, newline="") as f:
        f.write(merged)
    return len(file_paths), len(merged)


class TxtMergeCard(ttk.Frame):
    """TXT 合并卡片。"""

    def __init__(self, master):
        super().__init__(master, padding=10)
        self.files = []

        tip = ttk.Label(
            self,
            text="按列表顺序合并多个文本文件，内容格式不变。可拖拽文件到下方列表。",
            foreground="#555555",
        )
        tip.pack(fill="x", pady=(0, 4))

        # 文件列表
        frame = ttk.Frame(self)
        frame.pack(fill="both", expand=True, pady=4)
        self.listbox = tk.Listbox(frame, selectmode="extended")
        self.listbox.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=self.listbox.yview)
        scrollbar.pack(side="right", fill="y")
        self.listbox.config(yscrollcommand=scrollbar.set)

        # 操作按钮
        btn_frame = ttk.Frame(self)
        btn_frame.pack(fill="x", pady=4)
        ttk.Button(btn_frame, text="添加文件", command=self.add_files).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="添加文件夹", command=self.add_folder).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="移除选中", command=self.remove_selected).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="清空列表", command=self.clear_list).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="上移", command=lambda: self.move(-1)).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="下移", command=lambda: self.move(1)).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="按文件名排序", command=self.sort_by_name).pack(side="left", padx=2)

        # 衔接方式
        opt_frame = ttk.Frame(self)
        opt_frame.pack(fill="x", pady=2)
        ttk.Label(opt_frame, text="文件之间衔接方式：").pack(side="left")
        self.join_var = tk.StringVar(value="direct")
        ttk.Radiobutton(opt_frame, text="直接接在末尾", variable=self.join_var, value="direct").pack(side="left", padx=4)
        ttk.Radiobutton(opt_frame, text="新起一行再接", variable=self.join_var, value="newline").pack(side="left", padx=4)

        # 高级选项
        adv_frame = ttk.Frame(self)
        adv_frame.pack(fill="x", pady=2)
        self.header_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(adv_frame, text="文件名作为章节标题", variable=self.header_var).pack(side="left", padx=4)
        self.empty_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(adv_frame, text="去重空行", variable=self.empty_var).pack(side="left", padx=4)

        # 分隔符
        sep_frame = ttk.Frame(self)
        sep_frame.pack(fill="x", pady=2)
        ttk.Label(sep_frame, text="自定义分隔符（留空则不插入）：").pack(side="left")
        self.sep_var = tk.StringVar(value="")
        ttk.Entry(sep_frame, textvariable=self.sep_var, width=24).pack(side="left", padx=4)

        # 输出编码
        enc_frame = ttk.Frame(self)
        enc_frame.pack(fill="x", pady=2)
        ttk.Label(enc_frame, text="输出编码：").pack(side="left")
        self.enc_var = tk.StringVar(value="utf-8-sig")
        for enc, label in [("utf-8-sig", "UTF-8(带BOM)"), ("utf-8", "UTF-8"), ("gbk", "GBK")]:
            ttk.Radiobutton(enc_frame, text=label, variable=self.enc_var, value=enc).pack(side="left", padx=4)

        # 合并按钮
        merge_frame = ttk.Frame(self)
        merge_frame.pack(fill="x", pady=(6, 4))
        ttk.Button(merge_frame, text="合并为单个 txt", command=self.merge).pack(side="left", padx=2)

        self.status = ttk.Label(self, text="", foreground="#333333")
        self.status.pack(fill="x", pady=(0, 4))

        # 拖拽
        self._ole_drop = None
        if _HAS_DND:
            self.listbox.drop_target_register(DND_FILES)
            self.listbox.dnd_bind("<<Drop>>", self.on_drop)
        elif _HAS_OLE_DND:
            self._ole_drop = OleDropTarget(self.winfo_toplevel().winfo_id(), self._append_paths)
            if not self._ole_drop.register():
                self._ole_drop = None
                self.status.config(text="提示：拖拽功能初始化失败，可用按钮添加文件")
        else:
            self.status.config(text="提示：拖拽功能不可用，可用按钮添加文件")

    # ---- 文件操作 ----
    def add_files(self):
        paths = filedialog.askopenfilenames(
            title="选择要合并的文本文件（支持 txt / wnd 等任意文本格式）",
            filetypes=[("所有文件", "*.*"), ("文本文件", "*.txt")],
        )
        self._append_paths(paths)

    def add_folder(self):
        folder = filedialog.askdirectory(title="选择包含文本文件的文件夹")
        if not folder:
            return
        paths = [
            os.path.join(folder, name)
            for name in sorted(os.listdir(folder))
            if name.lower().endswith(".txt")
        ]
        if not paths:
            messagebox.showinfo("提示", "该文件夹内没有找到 .txt 文件。")
            return
        self._append_paths(paths)

    def on_drop(self, event):
        paths = self.winfo_toplevel().tk.splitlist(event.data)
        self._append_paths(paths)

    def _append_paths(self, paths):
        for p in paths:
            if os.path.isfile(p) and p not in self.files:
                self.files.append(p)
        self.refresh_list()

    def remove_selected(self):
        sel = list(self.listbox.curselection())
        for idx in reversed(sel):
            del self.files[idx]
        self.refresh_list()

    def clear_list(self):
        self.files = []
        self.refresh_list()

    def move(self, direction):
        sel = list(self.listbox.curselection())
        if not sel:
            return
        idx = sel[0]
        new_idx = idx + direction
        if new_idx < 0 or new_idx >= len(self.files):
            return
        self.files[idx], self.files[new_idx] = self.files[new_idx], self.files[idx]
        self.refresh_list()
        self.listbox.selection_set(new_idx)

    def sort_by_name(self):
        self.files.sort(key=lambda p: os.path.basename(p).lower())
        self.refresh_list()

    def refresh_list(self):
        self.listbox.delete(0, tk.END)
        for p in self.files:
            self.listbox.insert(tk.END, os.path.basename(p))
        self.status.config(text=f"共 {len(self.files)} 个文件")

    def merge(self):
        if not self.files:
            messagebox.showwarning("提示", "请先添加要合并的 txt 文件。")
            return
        output = filedialog.asksaveasfilename(
            title="保存合并后的文件",
            defaultextension=".txt",
            filetypes=[("文本文件", "*.txt")],
            initialfile="merged.txt",
        )
        if not output:
            return
        try:
            count, total_chars = merge_txt_files(
                self.files,
                output,
                newline_between=self.join_var.get() == "newline",
                add_headers=self.header_var.get(),
                separator=self.sep_var.get().strip(),
                remove_empty_lines=self.empty_var.get(),
                output_encoding=self.enc_var.get(),
            )
        except Exception as e:  # noqa: BLE001
            messagebox.showerror("合并失败", f"发生错误：\n{e}")
            return
        messagebox.showinfo("完成", f"已合并 {count} 个文件，共 {total_chars} 个字符。\n输出文件：\n{output}")
        self.status.config(text=f"合并完成：{output}")


# =====================================================================
# 卡片二：CSV 转换
# =====================================================================

def read_csv_content(path):
    """鲁棒读取 CSV 文件内容，返回 (编码, 分隔符, 字符串列表)。"""
    with open(path, "rb") as f:
        raw = f.read()
    raw_no_nul = raw.replace(b"\x00", b"")
    encodings = ["utf-8-sig", "utf-8", "gbk", "gb18030", "utf-16-le", "utf-16-be", "latin-1", "cp1252", "big5"]
    text = None
    used_enc = "utf-8"
    for enc in encodings:
        try:
            text = raw_no_nul.decode(enc)
            used_enc = enc
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if text is None:
        text = raw_no_nul.decode("utf-8", errors="replace")
        used_enc = "utf-8 (replace)"
    delim = ","
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=[",", ";", "\t", "|"])
        delim = dialect.delimiter
    except Exception:  # noqa: BLE001
        pass
    return used_enc, delim, text


def read_csv_rows(path, skip_header=False):
    """读取 CSV 内容，返回行列表（每行是单元格值列表）。

    skip_header=True 时跳过第一行（用于单工作表合并时只保留第一个文件的表头）。
    """
    _, delim, text = read_csv_content(path)
    rows = []
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    for r, row in enumerate(reader, start=1):
        if skip_header and r == 1:
            continue
        rows.append(list(row))
    return rows


def convert_csv_to_sheet(ws, path, skip_header=False):
    """把单个 CSV 内容写入给定工作表 ws，返回写入行数。"""
    rows = read_csv_rows(path, skip_header=skip_header)
    for r, row in enumerate(rows, start=1):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        width = 10
        for cell in ws[letter]:
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)) + 2, 60))
        ws.column_dimensions[letter].width = width
    return len(rows)


def safe_sheet_name(name, used):
    """生成合法且不重复的工作表名(<=31 字符)。"""
    base = name[:31]
    candidate = base
    i = 1
    while candidate in used:
        suffix = f"_{i}"
        candidate = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(candidate)
    return candidate


def csv_to_txt(path, out_path, output_encoding="utf-8-sig"):
    """把单个 CSV 转为 TXT（保留原始文本内容，仅去除 NUL 空字节）。"""
    _, _, text = read_csv_content(path)
    with open(out_path, "w", encoding=output_encoding, newline="") as f:
        f.write(text)
    return len(text.splitlines())


class CsvConvertCard(ttk.Frame):
    """CSV 转换卡片（转 XLSX / 转 TXT）。"""

    def __init__(self, master):
        super().__init__(master, padding=10)
        self.csv_files = []
        self.out_dir = ""

        # 文件选择
        frm = ttk.Frame(self)
        frm.pack(fill="x", pady=2)
        ttk.Label(frm, text="CSV 文件 / 文件夹：").pack(side="left")
        self.var_files = tk.StringVar(value="未选择")
        ttk.Label(frm, textvariable=self.var_files, foreground="blue").pack(side="left", padx=6)
        ttk.Button(frm, text="选择文件", command=self.pick_files).pack(side="left", padx=2)
        ttk.Button(frm, text="选择文件夹", command=self.pick_folder).pack(side="left", padx=2)

        # 输出目录
        frm2 = ttk.Frame(self)
        frm2.pack(fill="x", pady=2)
        ttk.Label(frm2, text="输出目录：").pack(side="left")
        self.var_out = tk.StringVar(value="与源文件同目录")
        ttk.Label(frm2, textvariable=self.var_out, foreground="blue").pack(side="left", padx=6)
        ttk.Button(frm2, text="选择输出目录", command=self.pick_out).pack(side="left", padx=2)

        # 转换类型
        frm3 = ttk.Frame(self)
        frm3.pack(fill="x", pady=2)
        ttk.Label(frm3, text="转换类型：").pack(side="left")
        self.conv_var = tk.StringVar(value="xlsx")
        ttk.Radiobutton(frm3, text="转 XLSX", variable=self.conv_var, value="xlsx").pack(side="left", padx=4)
        ttk.Radiobutton(frm3, text="转 TXT", variable=self.conv_var, value="txt").pack(side="left", padx=4)

        # 合并选项（仅 XLSX 有效）
        frm_merge = ttk.Frame(self)
        frm_merge.pack(fill="x", pady=2)
        ttk.Label(frm_merge, text="XLSX 输出方式：").pack(side="left")
        self.merge_var = tk.StringVar(value="single")
        ttk.Radiobutton(frm_merge, text="每个文件单独一个工作簿", variable=self.merge_var, value="single").pack(side="left", padx=4)
        ttk.Radiobutton(frm_merge, text="合并为多工作表", variable=self.merge_var, value="multi").pack(side="left", padx=4)
        ttk.Radiobutton(frm_merge, text="合并为单工作表", variable=self.merge_var, value="one").pack(side="left", padx=4)

        # 单工作表合并选项（仅合并为单工作表时有效）
        self.merge_single_header = tk.BooleanVar(value=True)
        self.merge_header_chk = ttk.Checkbutton(
            self, text="合并为单工作表时仅保留第一个文件的表头", variable=self.merge_single_header
        )
        self.merge_header_chk.pack(fill="x", pady=2)

        # 输出编码（仅 TXT 有效）
        frm4 = ttk.Frame(self)
        frm4.pack(fill="x", pady=2)
        ttk.Label(frm4, text="TXT 输出编码：").pack(side="left")
        self.txt_enc_var = tk.StringVar(value="utf-8-sig")
        for enc, label in [("utf-8-sig", "UTF-8(带BOM)"), ("utf-8", "UTF-8"), ("gbk", "GBK")]:
            ttk.Radiobutton(frm4, text=label, variable=self.txt_enc_var, value=enc).pack(side="left", padx=4)

        # 开始按钮
        frm5 = ttk.Frame(self)
        frm5.pack(fill="x", pady=4)
        self.btn_run = ttk.Button(frm5, text="开始转换", command=self.run)
        self.btn_run.pack(side="left", padx=2)

        # 进度条
        self.progress = ttk.Progressbar(self, mode="determinate")
        self.progress.pack(fill="x", pady=4)

        # 日志
        self.log_text = tk.Text(self, height=12, state="disabled")
        self.log_text.pack(fill="both", expand=True, pady=4)

    def log(self, msg):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")
        self.update_idletasks()

    def pick_files(self):
        files = filedialog.askopenfilenames(
            title="选择 CSV 文件", filetypes=[("CSV 文件", "*.csv"), ("全部", "*.*")]
        )
        if files:
            self.csv_files = list(files)
            self.var_files.set(f"已选 {len(self.csv_files)} 个文件")

    def pick_folder(self):
        folder = filedialog.askdirectory(title="选择包含 CSV 的文件夹")
        if folder:
            self.csv_files = [
                os.path.join(folder, f)
                for f in os.listdir(folder)
                if f.lower().endswith(".csv")
            ]
            self.var_files.set(f"文件夹: {folder} ({len(self.csv_files)} 个 CSV)")

    def pick_out(self):
        d = filedialog.askdirectory(title="选择输出目录")
        if d:
            self.out_dir = d
            self.var_out.set(d)

    def run(self):
        if not self.csv_files:
            messagebox.showwarning("提示", "请先选择 CSV 文件或文件夹。")
            return
        out_dir = self.out_dir or os.path.dirname(self.csv_files[0]) or "."
        os.makedirs(out_dir, exist_ok=True)

        self.btn_run.configure(state="disabled")
        self.progress["value"] = 0
        self.log("开始转换...")

        def progress(cur, total):
            self.progress["maximum"] = total
            self.progress["value"] = cur
            self.update_idletasks()

        conv = self.conv_var.get()
        try:
            if conv == "xlsx":
                merge_mode = self.merge_var.get()
                if merge_mode == "single":
                    self._build_single_xlsx(out_dir, progress)
                elif merge_mode == "multi":
                    self._build_merged_multi_sheet(out_dir, progress)
                else:
                    self._build_merged_single_sheet(out_dir, progress)
            else:
                self._build_txt(out_dir, progress)
            self.log("全部处理完成。")
            messagebox.showinfo("完成", "转换完成！")
        except Exception as e:  # noqa: BLE001
            self.log(f"发生错误: {e}")
            messagebox.showerror("错误", str(e))
        finally:
            self.btn_run.configure(state="normal")

    def _build_single_xlsx(self, out_dir, progress):
        total = len(self.csv_files)
        for idx, src in enumerate(self.csv_files, start=1):
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Sheet1"
                name = os.path.splitext(os.path.basename(src))[0]
                rows = convert_csv_to_sheet(ws, src)
                dst = os.path.join(out_dir, name + ".xlsx")
                wb.save(dst)
                self.log(f"[成功] {os.path.basename(src)} -> {os.path.basename(dst)} ({rows} 行)")
            except Exception as e:  # noqa: BLE001
                self.log(f"[失败] {os.path.basename(src)} 错误: {e}")
            progress(idx, total)

    def _build_merged_multi_sheet(self, out_dir, progress):
        """合并为单个工作簿，每个 CSV 一个工作表。"""
        wb = Workbook()
        wb.remove(wb.active)
        used = set()
        total = len(self.csv_files)
        for idx, src in enumerate(self.csv_files, start=1):
            try:
                name = safe_sheet_name(os.path.splitext(os.path.basename(src))[0], used)
                ws = wb.create_sheet(title=name)
                rows = convert_csv_to_sheet(ws, src)
                self.log(f"[成功] {os.path.basename(src)} -> 工作表[{name}] ({rows} 行)")
            except Exception as e:  # noqa: BLE001
                self.log(f"[失败] {os.path.basename(src)} 错误: {e}")
            progress(idx, total)
        if wb.sheetnames:
            dst = os.path.join(out_dir, "merged.xlsx")
            wb.save(dst)
            self.log(f"已合并保存到 {dst}")
        else:
            self.log("没有可转换的文件。")

    def _build_merged_single_sheet(self, out_dir, progress):
        """合并为单个工作簿，所有 CSV 纵向堆叠到同一个工作表。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "合并数据"
        only_first_header = self.merge_single_header.get()
        total = len(self.csv_files)
        current_row = 1
        for idx, src in enumerate(self.csv_files, start=1):
            try:
                skip = only_first_header and idx > 1
                rows = read_csv_rows(src, skip_header=skip)
                for row in rows:
                    for c, value in enumerate(row, start=1):
                        ws.cell(row=current_row, column=c, value=value)
                    current_row += 1
                self.log(f"[成功] {os.path.basename(src)} -> 追加 {len(rows)} 行")
            except Exception as e:  # noqa: BLE001
                self.log(f"[失败] {os.path.basename(src)} 错误: {e}")
            progress(idx, total)
        # 自动列宽
        for c in range(1, ws.max_column + 1):
            letter = get_column_letter(c)
            width = 10
            for cell in ws[letter]:
                if cell.value is not None:
                    width = max(width, min(len(str(cell.value)) + 2, 60))
            ws.column_dimensions[letter].width = width
        if current_row > 1:
            dst = os.path.join(out_dir, "merged_single.xlsx")
            wb.save(dst)
            self.log(f"已合并保存到 {dst}")
        else:
            self.log("没有可转换的文件。")

    def _build_txt(self, out_dir, progress):
        enc = self.txt_enc_var.get()
        total = len(self.csv_files)
        for idx, src in enumerate(self.csv_files, start=1):
            try:
                name = os.path.splitext(os.path.basename(src))[0]
                dst = os.path.join(out_dir, name + ".txt")
                rows = csv_to_txt(src, dst, output_encoding=enc)
                self.log(f"[成功] {os.path.basename(src)} -> {os.path.basename(dst)} ({rows} 行)")
            except Exception as e:  # noqa: BLE001
                self.log(f"[失败] {os.path.basename(src)} 错误: {e}")
            progress(idx, total)


# =====================================================================
# 卡片三：文本工具
# =====================================================================

def split_file_by_lines(path, out_dir, lines_per_file, output_encoding="utf-8-sig"):
    """按行数拆分文本文件为多个小文件，返回生成的文件列表。"""
    content, _ = read_text_file(path)
    all_lines = content.split("\n")
    base = os.path.splitext(os.path.basename(path))[0]
    created = []
    part = 1
    for i in range(0, len(all_lines), lines_per_file):
        chunk = all_lines[i:i + lines_per_file]
        dst = os.path.join(out_dir, f"{base}_part{part}.txt")
        with open(dst, "w", encoding=output_encoding, newline="") as f:
            f.write("\n".join(chunk))
        created.append(dst)
        part += 1
    return created


def dedupe_lines(path, out_path, output_encoding="utf-8-sig"):
    """按行去重，保留首次出现的行，返回 (原行数, 去重后行数)。"""
    content, _ = read_text_file(path)
    lines = content.split("\n")
    seen = set()
    unique = []
    for line in lines:
        if line not in seen:
            seen.add(line)
            unique.append(line)
    with open(out_path, "w", encoding=output_encoding, newline="") as f:
        f.write("\n".join(unique))
    return len(lines), len(unique)


def filter_lines(path, out_path, keyword, mode="keep", output_encoding="utf-8-sig"):
    """按关键词过滤行。mode='keep' 保留含关键词的行，mode='remove' 删除含关键词的行。"""
    content, _ = read_text_file(path)
    lines = content.split("\n")
    if mode == "keep":
        result = [line for line in lines if keyword in line]
    else:
        result = [line for line in lines if keyword not in line]
    with open(out_path, "w", encoding=output_encoding, newline="") as f:
        f.write("\n".join(result))
    return len(lines), len(result)


def replace_keyword(path, out_path, old, new, output_encoding="utf-8-sig"):
    """全文替换关键词，返回替换次数。"""
    content, _ = read_text_file(path)
    count = content.count(old)
    content = content.replace(old, new)
    with open(out_path, "w", encoding=output_encoding, newline="") as f:
        f.write(content)
    return count


def add_prefix_suffix(path, out_path, prefix="", suffix="", output_encoding="utf-8-sig"):
    """给每行加前缀/后缀，返回行数。"""
    content, _ = read_text_file(path)
    lines = content.split("\n")
    result = [prefix + line + suffix for line in lines]
    with open(out_path, "w", encoding=output_encoding, newline="") as f:
        f.write("\n".join(result))
    return len(result)


def convert_case(path, out_path, mode="upper", output_encoding="utf-8-sig"):
    """大小写转换。mode: upper/lower/title。"""
    content, _ = read_text_file(path)
    if mode == "upper":
        result = content.upper()
    elif mode == "lower":
        result = content.lower()
    else:
        result = content.title()
    with open(out_path, "w", encoding=output_encoding, newline="") as f:
        f.write(result)
    return len(result.splitlines())


def convert_encoding_file(path, out_path, target_encoding="utf-8-sig"):
    """文件编码转换（自动识别源编码，转为目标编码）。"""
    content, _ = read_text_file(path)
    with open(out_path, "w", encoding=target_encoding, newline="") as f:
        f.write(content)
    return len(content.splitlines())


class TextToolsCard(ttk.Frame):
    """文本工具卡片：拆分/去重/过滤/替换/前后缀/大小写/编码转换。"""

    def __init__(self, master):
        super().__init__(master, padding=10)
        self.files = []
        self.out_dir = ""

        # 文件选择
        frm = ttk.Frame(self)
        frm.pack(fill="x", pady=2)
        ttk.Label(frm, text="文本文件：").pack(side="left")
        self.var_files = tk.StringVar(value="未选择")
        ttk.Label(frm, textvariable=self.var_files, foreground="blue").pack(side="left", padx=6)
        ttk.Button(frm, text="选择文件", command=self.pick_files).pack(side="left", padx=2)
        ttk.Button(frm, text="选择文件夹", command=self.pick_folder).pack(side="left", padx=2)

        # 输出目录
        frm2 = ttk.Frame(self)
        frm2.pack(fill="x", pady=2)
        ttk.Label(frm2, text="输出目录：").pack(side="left")
        self.var_out = tk.StringVar(value="与源文件同目录")
        ttk.Label(frm2, textvariable=self.var_out, foreground="blue").pack(side="left", padx=6)
        ttk.Button(frm2, text="选择输出目录", command=self.pick_out).pack(side="left", padx=2)

        # 操作类型
        frm3 = ttk.Frame(self)
        frm3.pack(fill="x", pady=2)
        ttk.Label(frm3, text="操作类型：").pack(side="left")
        self.op_var = tk.StringVar(value="dedupe")
        ops = [
            ("dedupe", "按行去重"),
            ("split", "按行拆分"),
            ("filter", "关键词过滤"),
            ("replace", "关键词替换"),
            ("affix", "加前后缀"),
            ("case", "大小写转换"),
            ("encoding", "编码转换"),
        ]
        for val, label in ops:
            ttk.Radiobutton(frm3, text=label, variable=self.op_var, value=val).pack(side="left", padx=3)

        # 参数区
        self.param_frame = ttk.LabelFrame(self, text="参数设置")
        self.param_frame.pack(fill="x", pady=4)

        # 拆分行数
        self.split_frame = ttk.Frame(self.param_frame)
        ttk.Label(self.split_frame, text="每个文件行数：").pack(side="left")
        self.split_var = tk.StringVar(value="1000")
        ttk.Entry(self.split_frame, textvariable=self.split_var, width=10).pack(side="left", padx=4)

        # 过滤模式
        self.filter_frame = ttk.Frame(self.param_frame)
        ttk.Label(self.filter_frame, text="关键词：").pack(side="left")
        self.filter_kw = tk.StringVar(value="")
        ttk.Entry(self.filter_frame, textvariable=self.filter_kw, width=20).pack(side="left", padx=4)
        self.filter_mode = tk.StringVar(value="keep")
        ttk.Radiobutton(self.filter_frame, text="保留含关键词", variable=self.filter_mode, value="keep").pack(side="left", padx=2)
        ttk.Radiobutton(self.filter_frame, text="删除含关键词", variable=self.filter_mode, value="remove").pack(side="left", padx=2)

        # 替换
        self.replace_frame = ttk.Frame(self.param_frame)
        ttk.Label(self.replace_frame, text="查找：").pack(side="left")
        self.replace_old = tk.StringVar(value="")
        ttk.Entry(self.replace_frame, textvariable=self.replace_old, width=15).pack(side="left", padx=4)
        ttk.Label(self.replace_frame, text="替换为：").pack(side="left")
        self.replace_new = tk.StringVar(value="")
        ttk.Entry(self.replace_frame, textvariable=self.replace_new, width=15).pack(side="left", padx=4)

        # 前后缀
        self.affix_frame = ttk.Frame(self.param_frame)
        ttk.Label(self.affix_frame, text="前缀：").pack(side="left")
        self.affix_pre = tk.StringVar(value="")
        ttk.Entry(self.affix_frame, textvariable=self.affix_pre, width=12).pack(side="left", padx=4)
        ttk.Label(self.affix_frame, text="后缀：").pack(side="left")
        self.affix_suf = tk.StringVar(value="")
        ttk.Entry(self.affix_frame, textvariable=self.affix_suf, width=12).pack(side="left", padx=4)

        # 大小写
        self.case_frame = ttk.Frame(self.param_frame)
        self.case_mode = tk.StringVar(value="upper")
        ttk.Radiobutton(self.case_frame, text="大写", variable=self.case_mode, value="upper").pack(side="left", padx=2)
        ttk.Radiobutton(self.case_frame, text="小写", variable=self.case_mode, value="lower").pack(side="left", padx=2)
        ttk.Radiobutton(self.case_frame, text="首字母大写", variable=self.case_mode, value="title").pack(side="left", padx=2)

        # 编码转换
        self.encoding_frame = ttk.Frame(self.param_frame)
        ttk.Label(self.encoding_frame, text="目标编码：").pack(side="left")
        self.enc_target = tk.StringVar(value="utf-8-sig")
        for enc, label in [("utf-8-sig", "UTF-8(带BOM)"), ("utf-8", "UTF-8"), ("gbk", "GBK")]:
            ttk.Radiobutton(self.encoding_frame, text=label, variable=self.enc_target, value=enc).pack(side="left", padx=4)

        # 输出编码（通用）
        self.out_enc_frame = ttk.Frame(self.param_frame)
        ttk.Label(self.out_enc_frame, text="输出编码：").pack(side="left")
        self.out_enc = tk.StringVar(value="utf-8-sig")
        for enc, label in [("utf-8-sig", "UTF-8(带BOM)"), ("utf-8", "UTF-8"), ("gbk", "GBK")]:
            ttk.Radiobutton(self.out_enc_frame, text=label, variable=self.out_enc, value=enc).pack(side="left", padx=4)

        # 默认显示去重参数
        self._show_params("dedupe")

        # 开始按钮
        frm5 = ttk.Frame(self)
        frm5.pack(fill="x", pady=4)
        self.btn_run = ttk.Button(frm5, text="开始处理", command=self.run)
        self.btn_run.pack(side="left", padx=2)

        # 进度条
        self.progress = ttk.Progressbar(self, mode="determinate")
        self.progress.pack(fill="x", pady=4)

        # 日志
        self.log_text = tk.Text(self, height=10, state="disabled")
        self.log_text.pack(fill="both", expand=True, pady=4)

        # 操作类型切换时更新参数区
        self.op_var.trace_add("write", lambda *a: self._show_params(self.op_var.get()))

    def _show_params(self, op):
        """根据操作类型显示对应参数区。"""
        for frame in [self.split_frame, self.filter_frame, self.replace_frame,
                      self.affix_frame, self.case_frame, self.encoding_frame, self.out_enc_frame]:
            frame.pack_forget()
        if op == "split":
            self.split_frame.pack(fill="x", pady=2)
            self.out_enc_frame.pack(fill="x", pady=2)
        elif op == "filter":
            self.filter_frame.pack(fill="x", pady=2)
            self.out_enc_frame.pack(fill="x", pady=2)
        elif op == "replace":
            self.replace_frame.pack(fill="x", pady=2)
            self.out_enc_frame.pack(fill="x", pady=2)
        elif op == "affix":
            self.affix_frame.pack(fill="x", pady=2)
            self.out_enc_frame.pack(fill="x", pady=2)
        elif op == "case":
            self.case_frame.pack(fill="x", pady=2)
            self.out_enc_frame.pack(fill="x", pady=2)
        elif op == "encoding":
            self.encoding_frame.pack(fill="x", pady=2)
        else:  # dedupe
            self.out_enc_frame.pack(fill="x", pady=2)

    def log(self, msg):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")
        self.update_idletasks()

    def pick_files(self):
        files = filedialog.askopenfilenames(
            title="选择文本文件", filetypes=[("所有文件", "*.*"), ("文本文件", "*.txt")]
        )
        if files:
            self.files = list(files)
            self.var_files.set(f"已选 {len(self.files)} 个文件")

    def pick_folder(self):
        folder = filedialog.askdirectory(title="选择包含文本文件的文件夹")
        if folder:
            self.files = [
                os.path.join(folder, f)
                for f in os.listdir(folder)
                if f.lower().endswith((".txt", ".csv", ".log"))
            ]
            self.var_files.set(f"文件夹: {folder} ({len(self.files)} 个文件)")

    def pick_out(self):
        d = filedialog.askdirectory(title="选择输出目录")
        if d:
            self.out_dir = d
            self.var_out.set(d)

    def run(self):
        if not self.files:
            messagebox.showwarning("提示", "请先选择文本文件。")
            return
        out_dir = self.out_dir or os.path.dirname(self.files[0]) or "."
        os.makedirs(out_dir, exist_ok=True)

        self.btn_run.configure(state="disabled")
        self.progress["value"] = 0
        self.log("开始处理...")

        def progress(cur, total):
            self.progress["maximum"] = total
            self.progress["value"] = cur
            self.update_idletasks()

        op = self.op_var.get()
        enc = self.out_enc.get()
        total = len(self.files)
        try:
            for idx, src in enumerate(self.files, start=1):
                try:
                    self._process_one(src, out_dir, op, enc)
                except Exception as e:  # noqa: BLE001
                    self.log(f"[失败] {os.path.basename(src)} 错误: {e}")
                progress(idx, total)
            self.log("全部处理完成。")
            messagebox.showinfo("完成", "处理完成！")
        except Exception as e:  # noqa: BLE001
            self.log(f"发生错误: {e}")
            messagebox.showerror("错误", str(e))
        finally:
            self.btn_run.configure(state="normal")

    def _process_one(self, src, out_dir, op, enc):
        base = os.path.splitext(os.path.basename(src))[0]
        if op == "dedupe":
            dst = os.path.join(out_dir, base + "_去重.txt")
            orig, uniq = dedupe_lines(src, dst, output_encoding=enc)
            self.log(f"[去重] {os.path.basename(src)}: {orig} -> {uniq} 行 -> {os.path.basename(dst)}")
        elif op == "split":
            try:
                n = int(self.split_var.get())
            except ValueError:
                raise ValueError("每个文件行数必须是整数")
            created = split_file_by_lines(src, out_dir, n, output_encoding=enc)
            self.log(f"[拆分] {os.path.basename(src)} -> {len(created)} 个文件")
        elif op == "filter":
            kw = self.filter_kw.get()
            if not kw:
                raise ValueError("请输入过滤关键词")
            dst = os.path.join(out_dir, base + "_过滤.txt")
            mode = self.filter_mode.get()
            orig, kept = filter_lines(src, dst, kw, mode=mode, output_encoding=enc)
            self.log(f"[过滤] {os.path.basename(src)}: {orig} -> {kept} 行 -> {os.path.basename(dst)}")
        elif op == "replace":
            old = self.replace_old.get()
            if not old:
                raise ValueError("请输入要查找的内容")
            dst = os.path.join(out_dir, base + "_替换.txt")
            count = replace_keyword(src, dst, old, self.replace_new.get(), output_encoding=enc)
            self.log(f"[替换] {os.path.basename(src)}: 替换 {count} 处 -> {os.path.basename(dst)}")
        elif op == "affix":
            dst = os.path.join(out_dir, base + "_加前后缀.txt")
            rows = add_prefix_suffix(src, dst, self.affix_pre.get(), self.affix_suf.get(), output_encoding=enc)
            self.log(f"[前后缀] {os.path.basename(src)}: 处理 {rows} 行 -> {os.path.basename(dst)}")
        elif op == "case":
            dst = os.path.join(out_dir, base + "_大小写.txt")
            rows = convert_case(src, dst, self.case_mode.get(), output_encoding=enc)
            self.log(f"[大小写] {os.path.basename(src)}: 处理 {rows} 行 -> {os.path.basename(dst)}")
        elif op == "encoding":
            dst = os.path.join(out_dir, base + "_转码.txt")
            rows = convert_encoding_file(src, dst, target_encoding=self.enc_target.get())
            self.log(f"[编码转换] {os.path.basename(src)} -> {os.path.basename(dst)} ({rows} 行)")


# =====================================================================
# 主窗口
# =====================================================================

class App:
    def __init__(self, root):
        self.root = root
        root.title("文本处理工具")
        root.geometry("680x620")
        root.minsize(560, 480)

        # 三卡片切换
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill="both", expand=True, padx=8, pady=8)

        self.txt_card = TxtMergeCard(self.notebook)
        self.csv_card = CsvConvertCard(self.notebook)
        self.tools_card = TextToolsCard(self.notebook)

        self.notebook.add(self.txt_card, text="TXT 合并")
        self.notebook.add(self.csv_card, text="CSV 转换")
        self.notebook.add(self.tools_card, text="文本工具")


def main():
    if _HAS_DND:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()
    app = App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
