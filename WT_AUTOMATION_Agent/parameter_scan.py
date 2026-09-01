# encoding: utf-8
"""参数扫描（Parameter Scan）—— 从 Excel 参数表生成参数化流程定义。

实现多参数扫描规范（multi_parameter_scan.md Skill）的核心逻辑：
1. 读取 Excel/CSV 参数表（每行=一组参数）
2. 展开为多组参数化的步骤（stepParams 填入参数值）
3. 执行时 _resolve_dynamic_value 自动替换 ${stepParams.xxx} 模板

使用：：
    from WT_AUTOMATION_Agent.parameter_scan import ParameterScanner

    scanner = ParameterScanner()
    flow_def = scanner.scan(
        excel_path="params.xlsx",
        template_steps=base_steps,       # 模板步骤（含 ${stepParams.xxx} 占位）
        sheet_name="Sheet1",
        output_path="flow_scan_result.json",
    )
"""
from __future__ import annotations

import csv
import json
import os
from copy import deepcopy
from dataclasses import dataclass, field
from typing import Any


class StepModeFilterUnavailable(ValueError):
    """参数表含 stepMode 列且存在请求模式的行，但模板步骤全无 stepTags。

    stepMode 行级过滤无法生效。此时若继续展开会把所有模式的全套步骤塞给每行
    （如新建行混入复制步骤、复制链定位综合1失败），代价是运行期才暴露的错误。
    应中止启动而非静默全跑。
    """


# ---------------------------------------------------------------------------
# 数据模型
# ---------------------------------------------------------------------------

@dataclass
class ParameterRow:
    """一行参数数据。"""
    index: int
    """起始行号（Excel 中的起始行，2-based）。"""
    values: dict[str, str]
    """列头 → 值映射。"""

    def to_step_params(self) -> dict[str, str]:
        return dict(self.values)


@dataclass
class ScanConfig:
    """参数扫描配置。"""
    sheet_name: str = "Sheet1"
    """Excel Sheet 名称（仅 .xlsx 有效）。"""
    header_row: int = 1
    """表头所在行（1-based）。"""
    data_start_row: int = 2
    """数据起始行（1-based）。"""
    max_rows: int = 0
    """最大行数限制（0=不限制）。"""
    merge_mode: str = "duplicate"
    """合并模式：duplicate=每行参数生成一份完整步骤副本；inline=只替换 stepParams。"""


@dataclass
class ScanResult:
    """扫描结果。"""
    rows: list[ParameterRow] = field(default_factory=list)
    column_names: list[str] = field(default_factory=list)
    total_rows: int = 0
    source_path: str = ""


# ---------------------------------------------------------------------------
# 参数表读取器
# ---------------------------------------------------------------------------

def _snake_case(name: str) -> str:
    """简单中文/空格 → snake_case 规范化。"""
    import re
    name = name.strip()
    # 中文直接保留拼音首字母？不，当作通用键。用下划线替换非法字符
    name = re.sub(r'[\s\u3000]+', '_', name)
    name = re.sub(r'[^\w]', '_', name)
    # 去掉首尾下划线
    name = name.strip('_')
    # 全小写
    return name.lower()


class ParameterScanner:
    """参数扫描器：Excel/CSV → 参数化 flow_definition。

    支持格式：.xlsx / .xls / .csv / .tsv
    """

    # ------------------------------------------------------------------
    # Excel 读取
    # ------------------------------------------------------------------

    @staticmethod
    def read_excel(
        file_path: str,
        sheet_name: str = "Sheet1",
        header_row: int = 1,
        data_start_row: int = 2,
        max_rows: int = 0,
    ) -> ScanResult:
        """从 .xlsx/.xls 读取参数表。

        参数：
            file_path: Excel 文件路径
            sheet_name: 目标 Sheet 名
            header_row: 表头所在行（1-based）
            data_start_row: 数据起始行（1-based）
            max_rows: 最大读取行数（0=不限制）

        返回：
            ScanResult，含 column_names 和 rows 列表
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "需要 openpyxl 库。请执行: pip install openpyxl"
            ) from None

        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]

        # 读表头
        columns: list[str] = []
        header_cells = list(ws.iter_rows(
            min_row=header_row, max_row=header_row, values_only=True
        ))
        if header_cells:
            for cell in header_cells[0]:
                col = _snake_case(str(cell or ""))
                columns.append(col if col else f"col_{len(columns) + 1}")

        # 读数据行
        rows: list[ParameterRow] = []
        row_idx = data_start_row
        for data_cells in ws.iter_rows(
            min_row=data_start_row, values_only=True
        ):
            if max_rows > 0 and len(rows) >= max_rows:
                break
            if all(cell is None or str(cell).strip() == "" for cell in data_cells):
                row_idx += 1
                continue

            values: dict[str, str] = {}
            for ci, cell in enumerate(data_cells):
                if ci < len(columns):
                    values[columns[ci]] = str(cell).strip() if cell is not None else ""

            rows.append(ParameterRow(index=row_idx, values=values))
            row_idx += 1

        wb.close()
        return ScanResult(
            rows=rows,
            column_names=columns,
            total_rows=len(rows),
            source_path=file_path,
        )

    @staticmethod
    def read_csv(
        file_path: str,
        delimiter: str = ",",
        max_rows: int = 0,
    ) -> ScanResult:
        """从 .csv/.tsv 读取参数表。"""
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")

        if delimiter == "auto":
            ext = os.path.splitext(file_path)[1].lower()
            delimiter = "\t" if ext in (".tsv", ".tab") else ","

        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)
            header = next(reader)
            columns = [_snake_case(h) for h in header]

            rows: list[ParameterRow] = []
            for ri, data_row in enumerate(reader, start=2):
                if max_rows > 0 and len(rows) >= max_rows:
                    break
                if all(c.strip() == "" for c in data_row):
                    continue
                values: dict[str, str] = {}
                for ci, cell in enumerate(data_row):
                    if ci < len(columns):
                        values[columns[ci]] = str(cell).strip()
                rows.append(ParameterRow(index=ri, values=values))

        return ScanResult(
            rows=rows,
            column_names=columns,
            total_rows=len(rows),
            source_path=file_path,
        )

    # ------------------------------------------------------------------
    # 智能文件检测
    # ------------------------------------------------------------------

    @staticmethod
    def read_auto(
        file_path: str,
        sheet_name: str = "Sheet1",
        max_rows: int = 0,
    ) -> ScanResult:
        """自动检测文件类型读取参数表。"""
        ext = os.path.splitext(file_path)[1].lower()
        if ext in (".xlsx", ".xls", ".xlsm"):
            return ParameterScanner.read_excel(
                file_path, sheet_name=sheet_name, max_rows=max_rows
            )
        elif ext in (".csv", ".tsv", ".tab"):
            return ParameterScanner.read_csv(file_path, max_rows=max_rows)
        else:
            raise ValueError(f"不支持的参数表格式: {ext}")

    # ------------------------------------------------------------------
    # 参数扫描核心
    # ------------------------------------------------------------------

    @staticmethod
    def scan(
        excel_path: str,
        template_steps: list[dict[str, Any]],
        sheet_name: str = "Sheet1",
        output_path: str | None = None,
        max_rows: int = 0,
        tower_mode_override: str | None = None,
    ) -> dict[str, Any]:
        """核心扫描函数：读 Excel → 为每行参数复制模板步骤 → 注入 stepParams。

        参数：
            excel_path: 参数 Excel 文件路径
            template_steps: 模板步骤列表（含 ${stepParams.xxx} 占位符）
            sheet_name: Excel Sheet 名称
            output_path: 保存路径（可选）
            max_rows: 最大行数（0=全部）
            tower_mode_override: 可选。非 None 时强制覆盖所有行的 towerMode 列
                （如 "single"/"multi"），用于运行期按项目条件自动决定单塔/多塔，
                不改动参数表文件。

        返回：
            完整的 flow_definition 字典，可直接保存为 .json 执行
        """
        result = ParameterScanner.read_auto(
            excel_path, sheet_name=sheet_name, max_rows=max_rows
        )

        if not result.rows:
            return {
                "description": f"参数扫描结果（空：{os.path.basename(excel_path)} 无数据行）",
                "steps": [],
                "runtimeConfig": {},
                "flowPackages": [],
            }

        # 运行期 towerMode 覆盖：强制把每一行的 towerMode 置为同一值
        # （如项目解析出多个测风塔 → multi），不改参数表文件。
        if tower_mode_override is not None:
            _override = str(tower_mode_override).strip().lower()
            if _override in ("single", "multi"):
                for _row in result.rows:
                    _row.values["towermode"] = _override
            else:
                print(
                    f"[parameter_scan] tower_mode_override='{tower_mode_override}' 非法，忽略覆盖"
                    f"（合法值: ['multi', 'single']）"
                )

        scanned_steps: list[dict[str, Any]] = []
        step_counter = 0

        # 行级步骤过滤（阶段1/2）：若参数表含 stepMode 列，则按行的 stepMode 过滤
        # 模板步骤；含 towerMode 列时，再按行的 towerMode 做第二维过滤（单塔/多塔）。
        # 模板步骤通过 stepTags 声明模式（create/copy/copyfull）、towerTags 声明塔模式
        # （single/multi）。空/缺 = 通用，所有行执行。无对应列时退化为旧行为（全步包含）。
        column_lower = [c.lower() for c in result.column_names]
        filter_enabled = "stepmode" in column_lower
        valid_modes = {"create", "copy", "copyfull"}
        tower_filter_enabled = "towermode" in column_lower
        valid_tower_modes = {"single", "multi"}

        # 硬失败：请求模式过滤但模板零 stepTags → stepMode 过滤无法生效。直接中止，
        # 避免"每行全跑全部模式步骤"的静默错排（如新建行混入复制链、复制链定位综合1
        # 失败；该错误在运行期才暴露，代价高）。编辑器保存流程会剥离 stepTags 是主因。
        if filter_enabled:
            requests_mode = any(
                str(row.values.get("stepmode", "")).strip().lower() in valid_modes
                for row in result.rows
            )
            any_step_tags = any(
                ParameterScanner._normalize_step_tags(step.get("stepTags"))
                for step in template_steps
                if isinstance(step, dict)
            )
            if requests_mode and not any_step_tags:
                raise StepModeFilterUnavailable(
                    "参数表含 stepMode 列且存在非空模式行，但模板步骤全无 stepTags 字段。"
                    "stepMode 行级过滤无法生效（编辑器保存流程会剥离 stepTags）。"
                    "请为模板步骤注入 stepTags 后再运行；流程文件绕过编辑器直接编辑 JSON。"
                )
        # towerMode 维度的同样硬失败保护
        if tower_filter_enabled:
            requests_tower = any(
                str(row.values.get("towermode", "")).strip().lower() in valid_tower_modes
                for row in result.rows
            )
            any_tower_tags = any(
                ParameterScanner._normalize_step_tags(step.get("towerTags"))
                for step in template_steps
                if isinstance(step, dict)
            )
            if requests_tower and not any_tower_tags:
                raise StepModeFilterUnavailable(
                    "参数表含 towerMode 列且存在非空塔模式行，但模板步骤全无 towerTags 字段。"
                    "towerMode 行级过滤无法生效。请为模板步骤注入 towerTags 后再运行。"
                )

        for param_row in result.rows:
            # 参数行分隔注释（仅非第一行时添加）
            if scanned_steps:
                scanned_steps.append({
                    "id": f"scan_separator_{param_row.index}",
                    "name": f"--- 参数行 #{param_row.index} ---",
                    "stage": "separator",
                    "strategy": "separator",
                    "actionType": "log",
                    "enabled": True,
                    "description": (
                        f"参数组 #{param_row.index}: "
                        + ", ".join(f"{k}={v}" for k, v in list(param_row.values.items())[:5])
                    ),
                    "windowTitle": "",
                    "controls": [],
                    "actionConfig": {
                        "action": "log",
                        "message": (
                            f"[参数扫描] 行 #{param_row.index}: "
                            + ", ".join(f"{k}={v}" for k, v in list(param_row.values.items())[:5])
                        ),
                        "timeoutSeconds": 0.1,
                        "waitBefore": 0.0,
                        "waitAfter": 0.0,
                    },
                    "stepParams": {},
                    "auxChecks": [],
                    "fallbacks": [],
                })

            # 本行模式：仅当启用过滤且模式合法时生效；否则视为不过滤（全步包含）
            if filter_enabled:
                row_mode = str(param_row.values.get("stepmode", "")).strip().lower()
                if row_mode not in valid_modes:
                    if row_mode:
                        print(
                            f"[parameter_scan] 参数表第 {param_row.index + 1} 行 "
                            f"stepMode='{row_mode}' 非法，按全步处理"
                            f"（合法值: {sorted(valid_modes)}）"
                        )
                    row_mode = None
            else:
                row_mode = None
            # 塔模式维度：无 towermode 列时默认"single"（历史流程均为单塔），
            # 这样单塔步骤(towerTags=single)照常执行、多塔组(towerTags=multi)被排除；
            # 值非法时视为不过滤（与 stepMode 非法退化一致）。
            if tower_filter_enabled:
                tower_mode = str(param_row.values.get("towermode", "")).strip().lower()
                if tower_mode not in valid_tower_modes:
                    if tower_mode:
                        print(
                            f"[parameter_scan] 参数表第 {param_row.index + 1} 行 "
                            f"towerMode='{tower_mode}' 非法，按全步处理"
                            f"（合法值: {sorted(valid_tower_modes)}）"
                        )
                    tower_mode = None
            else:
                tower_mode = "single"

            for template_step in template_steps:
                # 行级过滤（双维 AND）：某维未启用/值非法时该维不过滤；
                # 步骤该维标签为空=通用；否则须含本行对应模式值。
                if row_mode is not None:
                    step_tags = ParameterScanner._normalize_step_tags(
                        template_step.get("stepTags")
                    )
                    if step_tags and row_mode not in step_tags:
                        continue
                if tower_mode is not None:
                    tower_tags = ParameterScanner._normalize_step_tags(
                        template_step.get("towerTags")
                    )
                    if tower_tags and tower_mode not in tower_tags:
                        continue
                step_counter += 1
                new_step = deepcopy(template_step)

                # 生成唯一 ID
                original_id = new_step.get("id", f"step_{step_counter}")
                new_step["id"] = f"{original_id}_scan{param_row.index}_{step_counter}"

                # 注入参数：合并模板步骤自带默认 stepParams 与参数行值（行值覆盖同名键）。
                # 否则参数表缺列会把步骤默认值（如 defaultturbine/weathername/mettowername）
                # 冲掉，导致 ${stepParams.xxx} 占位符失效、输入字面量。
                merged_params = dict(template_step.get("stepParams") or {})
                merged_params.update(param_row.to_step_params())
                new_step["stepParams"] = merged_params

                # 更新描述
                orig_desc = new_step.get("description", "")
                new_step["description"] = (
                    f"[参数扫描 #{param_row.index}] {orig_desc}"
                    if orig_desc else f"参数扫描 #{param_row.index}"
                )

                scanned_steps.append(new_step)

        # 构建流程定义
        flow_definition: dict[str, Any] = {
            "description": (
                f"参数扫描结果 ({os.path.basename(excel_path)}, "
                f"{len(result.rows)}行 × {len(template_steps)}步 = {len(scanned_steps) - len(result.rows)}个有效步骤"
            ),
            "steps": scanned_steps,
            "runtimeConfig": {},
            "flowPackages": [],
        }

        if output_path:
            os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(flow_definition, f, ensure_ascii=False, indent=2)

        return flow_definition

    # ------------------------------------------------------------------
    # 快捷入口：从已有 flow_definition 扫描
    # ------------------------------------------------------------------

    @staticmethod
    def scan_from_flow(
        flow_json_path: str,
        excel_path: str,
        sheet_name: str = "Sheet1",
        output_path: str | None = None,
        max_rows: int = 0,
    ) -> dict[str, Any]:
        """从已有 flow_definition.json 读取模板步骤并执行参数扫描。"""
        with open(flow_json_path, "r", encoding="utf-8") as f:
            flow_def = json.load(f)
        template_steps = flow_def.get("steps", [])
        return ParameterScanner.scan(
            excel_path=excel_path,
            template_steps=template_steps,
            sheet_name=sheet_name,
            output_path=output_path,
            max_rows=max_rows,
        )

    @staticmethod
    def _normalize_step_tags(tags: Any) -> set[str]:
        """将 stepTags 规整为小写模式集合。

        stepTags 可为：None / []（通用，返回空集）、字符串（逗号分隔）、列表。
        模式取值：create / copy / copyfull（详见 scan 行级过滤逻辑）。
        """
        if not tags:
            return set()
        if isinstance(tags, str):
            tags = [t for t in tags.split(",") if t.strip()]
        result: set[str] = set()
        for t in tags:
            if isinstance(t, str):
                result.add(t.strip().lower())
        return result

    # ------------------------------------------------------------------
    # 步骤 Excel 分析 —— 发现可参数化字段（与 flow_excel_io.py 联动）
    # ------------------------------------------------------------------

    # 不可参数化的列（每步的元数据 / 固定标识）
    _NON_PARAM_COLUMNS: set = {
        "seq", "id", "name", "stage", "strategy", "actionType",
        "topLevel", "enabled", "packageRef", "codeSymbol", "codeReference",
        "controlId", "action", "fallbackMode", "fallbackTemplate",
        "regionAnchor", "inspectHintsJson", "actionConfigJson",
        "auxChecksJson", "fallbacksJson",
    }

    @classmethod
    def analyze_step_excel(
        cls,
        step_excel_path: str,
        sheet_name: str = "steps",
    ) -> dict[str, Any]:
        """分析 flow_steps.xlsx（flow_excel_io 导出的步骤表），
        发现哪些字段在不同步骤间存在差异（可参数化候选）。

        参数：
            step_excel_path: flow_excel_io.py 导出的步骤 Excel 路径
            sheet_name: 步骤数据所在的 sheet 名（默认 "steps"）

        返回：
            {
                "total_steps": N,
                "total_columns": N,
                "variant_columns": {列名: {"count": 唯一值数, "values": [...前10个值]}},
                "fixed_columns": [列名列表],
                "suggested_params": [建议作为参数的列名列表],
            }
        """
        import openpyxl

        if not os.path.isfile(step_excel_path):
            raise FileNotFoundError(f"步骤 Excel 不存在: {step_excel_path}")

        wb = openpyxl.load_workbook(step_excel_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{sheet_name}' 不存在，可用: {wb.sheetnames}")

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 2:
            return {"total_steps": 0, "total_columns": 0, "variant_columns": {},
                    "fixed_columns": [], "suggested_params": []}

        # 第一行 = 表头
        headers = [str(h or "").strip() for h in rows[0]]
        data_rows = rows[1:]

        # 按列收集所有值
        col_values: dict[str, list[str]] = {h: [] for h in headers}
        for row in data_rows:
            for ci, cell in enumerate(row):
                if ci < len(headers):
                    col_values[headers[ci]].append(str(cell or "").strip())

        variant_columns: dict[str, dict] = {}
        fixed_columns: list[str] = []
        suggested_params: list[str] = []

        for col_name, values in col_values.items():
            # 排除空列
            non_empty = [v for v in values if v]
            if not non_empty:
                fixed_columns.append(col_name)
                continue

            distinct = sorted(set(non_empty))
            if len(distinct) == 1:
                fixed_columns.append(col_name)
            else:
                variant_columns[col_name] = {
                    "count": len(distinct),
                    "values": distinct[:10],
                }
                if col_name not in cls._NON_PARAM_COLUMNS:
                    suggested_params.append(col_name)

        return {
            "total_steps": len(data_rows),
            "total_columns": len(headers),
            "variant_columns": variant_columns,
            "fixed_columns": fixed_columns,
            "suggested_params": suggested_params,
        }

    @classmethod
    def suggest_param_columns(
        cls,
        step_excel_path: str,
        sheet_name: str = "steps",
    ) -> list[str]:
        """快捷方法：返回建议的参数列名列表。"""
        analysis = cls.analyze_step_excel(step_excel_path, sheet_name)
        return analysis.get("suggested_params", [])

    @classmethod
    def export_param_template(
        cls,
        step_excel_path: str,
        output_path: str,
        selected_columns: list[str] | None = None,
        sheet_name: str = "steps",
        max_rows_hint: int = 3,
    ) -> str:
        """从步骤 Excel 生成参数模板 Excel（用户填入多组参数值即可扫描）。

        参数：
            step_excel_path: flow_excel_io 导出的步骤表
            output_path: 输出的参数模板 Excel 路径
            selected_columns: 要纳入模板的列（None=自动建议所有可变字段）
            sheet_name: 步骤 Sheet 名
            max_rows_hint: 模板预填行数（建议值，用户可增减）

        返回：
            输出文件路径
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        if selected_columns is None:
            selected_columns = cls.suggest_param_columns(step_excel_path, sheet_name)

        if not selected_columns:
            raise ValueError("未发现可参数化的字段，请检查步骤 Excel 或手动指定 selected_columns")

        # 读取步骤表获取当前值作为"默认值"
        wb_src = openpyxl.load_workbook(step_excel_path, read_only=True, data_only=True)
        ws_src = wb_src[sheet_name]
        src_rows = list(ws_src.iter_rows(values_only=True))
        wb_src.close()

        headers = [str(h or "").strip() for h in src_rows[0]]
        # 取第一行数据作为"默认值"参考
        default_row = src_rows[1] if len(src_rows) > 1 else src_rows[0]

        # 创建参数模板工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "参数表"

        # 样式
        header_font = Font(name="Microsoft YaHei UI", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hint_font = Font(name="Microsoft YaHei UI", color="888888", size=9, italic=True)
        border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        # 写入表头
        for ci, col in enumerate(selected_columns, start=1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # 写入提示行（第2行 = 说明）
        for ci, col in enumerate(selected_columns, start=1):
            col_idx = headers.index(col) if col in headers else -1
            default_val = str(default_row[col_idx]) if col_idx >= 0 and col_idx < len(default_row) else ""
            hint_text = f"默认值: {default_val}" if default_val else "待填写"
            cell = ws.cell(row=2, column=ci, value=hint_text)
            cell.font = hint_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # 预填几行空白 + 默认值
        for ri in range(max_rows_hint):
            row_num = ri + 3
            for ci, col in enumerate(selected_columns, start=1):
                col_idx = headers.index(col) if col in headers else -1
                default_val = str(default_row[col_idx]) if col_idx >= 0 and col_idx < len(default_row) else ""
                cell = ws.cell(row=row_num, column=ci, value=default_val)
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

        # 自动列宽
        for ci, col in enumerate(selected_columns, start=1):
            max_len = max(len(col), 12)
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = min(max_len + 4, 40)

        # 添加说明 Sheet
        ws_guide = wb.create_sheet("使用说明")
        guide_lines = [
            "参数模板使用说明",
            "",
            "1. 第1行（蓝色表头）= 参数名，对应步骤中的 ${stepParams.xxx} 占位符",
            "2. 第2行（灰色）= 当前步骤中的默认值，仅供参考",
            "3. 从第3行开始逐行填写多组参数值",
            "4. 每行参数会驱动整套步骤执行一次（如 3 行 = 跑 3 轮）",
            "5. 将此文件保存后，用 parameter_scan.scan() 生成参数化流程",
            "",
            "生成来源: " + os.path.basename(step_excel_path),
            f"参数列: {', '.join(selected_columns)}",
        ]
        for ri, line in enumerate(guide_lines, start=1):
            ws_guide.cell(row=ri, column=1, value=line)

        wb.save(output_path)
        return output_path

    @classmethod
    def auto_scan_from_steps(
        cls,
        step_excel_path: str,
        param_excel_path: str,
        output_path: str | None = None,
        max_rows: int = 0,
    ) -> dict[str, Any]:
        """一站式智能扫描：从步骤 Excel 出发生成参数化流程。

        1. 分析步骤 Excel → 发现可参数化字段
        2. 读取参数 Excel → 获取多组参数值
        3. 以步骤表中已有步骤为模板 → 注入参数 → 生成扫描结果

        这是 parameter_scan ↔ flow_excel_io 的核心桥接点。
        """
        import openpyxl

        # 步骤 1：分析步骤表，获取模板步骤
        wb = openpyxl.load_workbook(step_excel_path, read_only=True, data_only=True)
        ws = wb["steps"] if "steps" in wb.sheetnames else wb[wb.sheetnames[0]]
        src_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(src_rows) < 2:
            raise ValueError("步骤 Excel 为空（无数据行）")

        headers = [str(h or "").strip() for h in src_rows[0]]
        data_rows = src_rows[1:]

        # 步骤 2：读取参数表
        param_result = cls.read_auto(param_excel_path, max_rows=max_rows)
        if not param_result.rows:
            raise ValueError("参数 Excel 为空（无数据行）")

        # 步骤 3：构建模板步骤（从步骤 Excel 中提取）
        template_steps: list[dict[str, Any]] = []
        for row_idx, row in enumerate(data_rows):
            step: dict[str, Any] = {
                "id": f"step_{row_idx + 1}",
                "name": "",
                "stage": "captured",
                "strategy": "action",
                "actionType": "action",
                "topLevel": True,
                "enabled": True,
                "packageRef": "",
                "codeSymbol": "",
                "codeReference": "",
                "windowTitle": "",
                "successLog": "",
                "description": "",
                "controls": [],
                "actionConfig": {},
                "stepParams": {},
                "auxChecks": [],
                "fallbacks": [],
                "inspectHints": {},
            }
            for ci, cell in enumerate(row):
                col = headers[ci] if ci < len(headers) else ""
                val = str(cell or "").strip()
                if not col or not val:
                    continue
                _apply_step_cell(step, col, val)
            template_steps.append(step)

        # 步骤 4：参数扫描
        return cls.scan(
            excel_path=param_excel_path,
            template_steps=template_steps,
            output_path=output_path,
            max_rows=max_rows,
        )


# ---------------------------------------------------------------------------
# 步骤 Excel 单元格 → step JSON 映射（与 flow_excel_io 列名对齐）
# ---------------------------------------------------------------------------

def _apply_step_cell(step: dict, col: str, val: str) -> None:
    """将步骤 Excel 的一列值映射回 step 字典的对应字段。"""
    # 直接字段
    if col == "name":
        step["name"] = val
    elif col == "stage":
        step["stage"] = val
    elif col == "strategy":
        step["strategy"] = val
    elif col == "actionType":
        step["actionType"] = val
    elif col == "windowTitle":
        step["windowTitle"] = val
    elif col == "successLog":
        step["successLog"] = val
    elif col == "description":
        step["description"] = val
    elif col == "packageRef":
        step["packageRef"] = val
    elif col == "onError":
        step.setdefault("actionConfig", {})["onError"] = val
    elif col == "fallbackTemplate":
        step.setdefault("actionConfig", {})["fallbackTemplate"] = val
    elif col == "fallbackMode":
        step.setdefault("actionConfig", {})["fallbackMode"] = val
    elif col == "saveAs":
        step.setdefault("actionConfig", {})["saveAs"] = val
    elif col == "retryCount":
        step.setdefault("actionConfig", {})["retryCount"] = int(val) if val else 0
    elif col == "retryInterval":
        step.setdefault("actionConfig", {})["retryInterval"] = float(val) if val else 1.0
    elif col == "action":
        step.setdefault("actionConfig", {})["action"] = val
    elif col == "controlId":
        step.setdefault("actionConfig", {})["controlId"] = val

    # 输入/文本（支持 ${stepParams.xxx} 模板语法）
    elif col == "inputText":
        step.setdefault("actionConfig", {})["text"] = val

    # 时间相关
    elif col == "timeoutSeconds":
        step.setdefault("actionConfig", {})["timeoutSeconds"] = float(val) if val else 3.0
    elif col == "waitBefore":
        step.setdefault("actionConfig", {})["waitBefore"] = float(val) if val else 0.0
    elif col == "waitAfter":
        step.setdefault("actionConfig", {})["waitAfter"] = float(val) if val else 0.3

    # 相对区域字段
    elif col in ("regionX", "regionY", "regionWidth", "regionHeight"):
        key = col.replace("region", "").lower()
        step.setdefault("actionConfig", {}).setdefault("relativeRegion", {})[key] = (
            float(val) if val else 0
        )
    elif col == "regionAnchor":
        step.setdefault("actionConfig", {}).setdefault("relativeRegion", {})["anchor"] = val

    # 父窗口字段
    elif col == "parentWindowTitle":
        step.setdefault("actionConfig", {}).setdefault("parentWindow", {})["title"] = val
    elif col == "parentWindowClassName":
        step.setdefault("actionConfig", {}).setdefault("parentWindow", {})["className"] = val
    elif col == "parentWindowFrameworkId":
        step.setdefault("actionConfig", {}).setdefault("parentWindow", {})["frameworkId"] = val

    # continueWhen 字段
    elif col == "continueWhenControlId":
        step.setdefault("actionConfig", {}).setdefault("continueWhen", {})["controlId"] = val
    elif col == "continueWhenCondition":
        step.setdefault("actionConfig", {}).setdefault("continueWhen", {})["condition"] = val
    elif col == "continueWhenTimeoutSeconds":
        step.setdefault("actionConfig", {}).setdefault("continueWhen", {})["timeoutSeconds"] = (
            float(val) if val else 0.0
        )
    elif col == "continueWhenWindowTitleHint":
        step.setdefault("actionConfig", {}).setdefault("continueWhen", {})["windowTitleHint"] = val

    # inspectHints 字段
    elif col == "inspectControlName":
        step.setdefault("inspectHints", {})["controlName"] = val
    elif col == "inspectClassName":
        step.setdefault("inspectHints", {})["className"] = val
    elif col == "inspectAutomationId":
        step.setdefault("inspectHints", {})["automationId"] = val
    elif col == "inspectControlType":
        step.setdefault("inspectHints", {})["controlType"] = val
    elif col == "inspectUiPath":
        step.setdefault("inspectHints", {})["uiPath"] = val
    elif col == "inspectTemplateKey":
        step.setdefault("inspectHints", {})["templateKey"] = val

    # JSON 字段
    elif col == "stepParamsJson" and val:
        try:
            step["stepParams"] = json.loads(val)
        except json.JSONDecodeError:
            step["stepParams"] = {}
    elif col == "actionConfigJson" and val:
        try:
            extra = json.loads(val)
            if isinstance(extra, dict):
                step.setdefault("actionConfig", {}).update(extra)
        except json.JSONDecodeError:
            pass
