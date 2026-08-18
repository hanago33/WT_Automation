import argparse
import json
import os
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError as exc:  # pragma: no cover
    Workbook = None
    load_workbook = None
    Alignment = None
    Font = None
    PatternFill = None
    Comment = None
    DataValidation = None
    get_column_letter = None
    OPENPYXL_IMPORT_ERROR = exc
else:
    OPENPYXL_IMPORT_ERROR = None

from wt_action_schema import (
    ALLOWED_CONTINUE_WHEN_CONDITIONS,
    ALLOWED_ON_ERROR_MODES,
    ALLOWED_PARENT_WINDOW_FRAMEWORK_IDS,
    ALLOWED_RELATIVE_REGION_ANCHORS,
    build_action_schema_hint,
    get_action_names,
    get_action_schema,
)


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_FLOW_JSON = os.path.join(BASE_DIR, "workspace", "flow_definition.json")
DEFAULT_FLOW_XLSX = os.path.join(BASE_DIR, "workspace", "flow_steps.xlsx")
OPTIONS_SHEET_NAME = "_options"
ACTION_GUIDE_SHEET_NAME = "action_guide"
EXAMPLES_SHEET_NAME = "examples"
ROUNDTRIP_AUDIT_BASELINE_PATH_FIELD = "audit.baselineFlowPath"
ROUNDTRIP_AUDIT_STEP_IDS_FIELD = "audit.stepIdsJson"
ROUNDTRIP_AUDIT_BASELINE_LABEL_FIELD = "audit.baselineLabel"

STEP_COLUMNS = [
    "seq",
    "id",
    "name",
    "stage",
    "strategy",
    "actionType",
    "topLevel",
    "enabled",
    "packageRef",
    "codeSymbol",
    "codeReference",
    "windowTitle",
    "successLog",
    "description",
    "notes",
    "inspectControlName",
    "inspectClassName",
    "inspectAutomationId",
    "inspectControlType",
    "inspectUiPath",
    "inspectTemplateKey",
    "action",
    "controlId",
    "inputText",
    "postInputKeys",
    "waitSeconds",
    "wheelDelta",
    "timeoutSeconds",
    "waitBefore",
    "waitAfter",
    "continueWhenControlId",
    "continueWhenCondition",
    "continueWhenTimeoutSeconds",
    "continueWhenWindowTitleHint",
    "retryCount",
    "retryInterval",
    "onError",
    "fallbackMode",
    "fallbackTemplate",
    "saveAs",
    "parentWindowTitle",
    "parentWindowClassName",
    "parentWindowFrameworkId",
    "regionX",
    "regionY",
    "regionWidth",
    "regionHeight",
    "regionAnchor",
    "inspectHintsJson",
    "stepParamsJson",
    "actionConfigJson",
    "auxChecksJson",
    "fallbacksJson",
    "fallbackChainJson",
]

CONTROL_COLUMNS = [
    "stepId",
    "order",
    "id",
    "name",
    "role",
    "targetMethod",
    "targetValue",
    "windowTitle",
    "templateKey",
    "targetIndex",
    "inspectDataJson",
    "auxChecksJson",
    "notes",
]

PACKAGE_COLUMNS = [
    "id",
    "name",
    "description",
    "stepIdsJson",
]

ACTION_TYPE_OPTIONS = ["script", "action", "flow_ref", "placeholder"]
STRATEGY_OPTIONS = ["script", "action", "flow_ref", "converted", "placeholder"]
BOOLEAN_OPTIONS = ["是", "否"]
ON_ERROR_OPTIONS = ["", *ALLOWED_ON_ERROR_MODES]
FALLBACK_MODE_OPTIONS = ["", "template_match"]
FRAMEWORK_OPTIONS = ["", *ALLOWED_PARENT_WINDOW_FRAMEWORK_IDS]
ANCHOR_OPTIONS = ["", *ALLOWED_RELATIVE_REGION_ANCHORS]
CONTINUE_CONDITION_OPTIONS = ["", *ALLOWED_CONTINUE_WHEN_CONDITIONS]
COMMON_TARGET_METHOD_OPTIONS = [
    "",
    "automation_id",
    "name",
    "class_name",
    "control_type",
    "name,class_name",
    "automation_id,control_type",
    "name,control_type",
    "class_name,control_type",
]
KNOWN_ACTION_CONFIG_KEYS = {
    "action",
    "controlId",
    "text",
    "value",
    "postInputKeys",
    "seconds",
    "delta",
    "timeoutSeconds",
    "waitBefore",
    "waitAfter",
    "continueWhen",
    "retryCount",
    "retryInterval",
    "onError",
    "fallbackMode",
    "fallbackTemplate",
    "saveAs",
    "parentWindow",
    "relativeRegion",
}
CORE_REQUIRED_STEP_COLUMNS = {"id", "name", "strategy", "actionType", "enabled"}
HIGH_FREQUENCY_ENGINEERING_COLUMNS = {
    "action",
    "controlId",
    "inputText",
    "postInputKeys",
    "timeoutSeconds",
    "waitBefore",
    "waitAfter",
    "continueWhenControlId",
    "continueWhenCondition",
    "continueWhenTimeoutSeconds",
    "retryCount",
    "retryInterval",
    "parentWindowTitle",
    "parentWindowClassName",
    "parentWindowFrameworkId",
    "regionX",
    "regionY",
    "regionWidth",
    "regionHeight",
}
STEP_COLUMN_COMMENTS = {
    "id": "步骤唯一标识，建议保持 step_xx 形式，导入时按此作为主键。",
    "name": "工程师可读名称，建议直接写业务动作，例如“点击-添加到数据”。",
    "strategy": "推荐使用 action 或 script。可从下拉中选择。",
    "actionType": "步骤类型。绝大多数自动化动作使用 action。",
    "enabled": "是否启用该步骤。是=参与运行，否=导入后保留但不执行。",
    "action": "动作名称。支持下拉，决定本步骤的主要必填列。",
    "controlId": "动作命中的目标控件 ID。若该动作依赖 controls 表中的控件，请填写这里。",
    "inputText": "用于输入类动作的文本内容，例如 type_text、type_text_relative。",
    "postInputKeys": "输入完成后追加发送的按键。常用值: {TAB} / {ENTER}；留空表示不补按键。",
    "timeoutSeconds": "当前步骤最大等待时间，单位秒。",
    "waitAfter": "动作完成后的固定等待时间。建议保留用于调试或微小缓冲，不建议完全依赖它做大等待。",
    "continueWhenControlId": "续跑条件目标控件 ID。填写后，步骤会在动作完成后等待该条件满足再继续。",
    "continueWhenCondition": "续跑条件类型。支持 exists / present / visible / enabled / gone。",
    "continueWhenTimeoutSeconds": "续跑条件最大等待时间，单位秒。",
    "continueWhenWindowTitleHint": "可选。若续跑条件要限定窗口标题，可填写这里辅助定位。",
    "parentWindowTitle": "父窗口相对区域动作的父窗口标题。",
    "parentWindowClassName": "父窗口相对区域动作的父窗口类名。",
    "parentWindowFrameworkId": "父窗口相对区域动作的框架类型，例如 WPF / Win32。",
    "regionX": "相对区域左上角 X 比例，范围通常在 0-1。",
    "regionY": "相对区域左上角 Y 比例，范围通常在 0-1。",
    "regionWidth": "相对区域宽度比例，范围通常在 0-1。",
    "regionHeight": "相对区域高度比例，范围通常在 0-1。",
    "inspectHintsJson": "仅补充显式列未覆盖的 inspectHints 字段；常规维护优先改显式列。",
    "actionConfigJson": "仅补充显式列未覆盖的 actionConfig 字段；常规维护优先改显式列。",
    "stepParamsJson": "用于保留脚本步骤或复杂动作的附加参数。",
}
CONTROL_COLUMN_COMMENTS = {
    "stepId": "所属步骤 ID，必须与 steps 表中的 id 对应。",
    "id": "控件唯一 ID，供 steps.action/controlId 引用。",
    "name": "控件可读名称。",
    "targetMethod": "控件定位方法，支持下拉组合。",
    "targetValue": "与 targetMethod 对应的值；多个参数使用英文逗号分隔。",
}
HEADER_FILL_CORE_REQUIRED = "FDE68A"
HEADER_FILL_HIGH_FREQUENCY = "BFDBFE"
HEADER_FILL_JSON_FALLBACK = "E9D5FF"
HEADER_FILL_OPTIONAL = "D9EAF7"


def _require_openpyxl():
    if Workbook is None or load_workbook is None:
        raise RuntimeError(
            "当前环境缺少 openpyxl，无法处理 Excel。请先执行: py -3.11 -m pip install openpyxl"
        ) from OPENPYXL_IMPORT_ERROR


def _read_json_file(file_path):
    with open(file_path, "r", encoding="utf-8") as file_obj:
        return json.load(file_obj)


def _write_json_file(file_path, payload):
    with open(file_path, "w", encoding="utf-8") as file_obj:
        json.dump(payload, file_obj, ensure_ascii=False, indent=2)


def _get_nested_dict_value(payload, dotted_path):
    current = payload
    for part in str(dotted_path or "").split("."):
        if not isinstance(current, dict):
            return None
        current = current.get(part)
    return current


def _normalize_audit_value(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    if value is None:
        return ""
    return str(value)


ROUNDTRIP_AUDIT_FIELD_SPECS = [
    {
        "field": "windowTitle",
        "label": "顶层窗口标题",
        "actions": {"click_relative_region", "type_text_relative"},
    },
    {
        "field": "actionConfig.parentWindow.title",
        "label": "父窗口标题",
        "actions": {"click_relative_region", "type_text_relative"},
    },
    {
        "field": "actionConfig.fallbackTemplate",
        "label": "模板兜底图片",
        "actions": {"click_relative_region", "type_text_relative"},
        "required_when": lambda step: _safe_text(
            _get_nested_dict_value(step, "actionConfig.fallbackMode")
        )
        == "template_match",
    },
    {
        "field": "actionConfig.continueWhen",
        "label": "续跑条件",
        "actions": None,
        "required_when": lambda step: isinstance(
            _get_nested_dict_value(step, "actionConfig.continueWhen"), dict
        )
        and bool(_get_nested_dict_value(step, "actionConfig.continueWhen")),
    },
]


def _build_step_map(payload):
    step_map = {}
    for step in (payload or {}).get("steps", []):
        if not isinstance(step, dict):
            continue
        step_id = _safe_text(step.get("id"))
        if step_id:
            step_map[step_id] = step
    return step_map


def load_flow_excel_meta(excel_path=DEFAULT_FLOW_XLSX):
    _require_openpyxl()
    workbook = load_workbook(excel_path, data_only=False)
    try:
        if "meta" not in workbook.sheetnames:
            return {}
        meta_rows = _sheet_to_dicts(workbook["meta"])
        return {
            _safe_text(row.get("field")): _safe_text(row.get("value"))
            for row in meta_rows
            if _safe_text(row.get("field"))
        }
    finally:
        try:
            workbook.close()
        except Exception:
            pass


def _parse_roundtrip_audit_step_ids(raw_value):
    parsed_value = _json_loads(raw_value, [])
    if not isinstance(parsed_value, list):
        return []
    step_ids = []
    seen_step_ids = set()
    for item in parsed_value:
        step_id = _safe_text(item)
        if step_id and step_id not in seen_step_ids:
            step_ids.append(step_id)
            seen_step_ids.add(step_id)
    return step_ids


def compare_roundtrip_fields_between_payloads(baseline_payload, candidate_payload, step_ids=None):
    baseline_steps = _build_step_map(baseline_payload or {})
    candidate_steps = _build_step_map(candidate_payload or {})
    ordered_step_ids = []
    if step_ids:
        seen = set()
        for step_id in step_ids:
            normalized = _safe_text(step_id)
            if normalized and normalized not in seen:
                ordered_step_ids.append(normalized)
                seen.add(normalized)
    else:
        ordered_step_ids = list(candidate_steps.keys())

    issues = []
    for step_id in ordered_step_ids:
        candidate_step = candidate_steps.get(step_id)
        if not isinstance(candidate_step, dict):
            continue
        baseline_step = baseline_steps.get(step_id)
        if not isinstance(baseline_step, dict):
            issues.append(
                {
                    "stepId": step_id,
                    "stepName": _safe_text(candidate_step.get("name")),
                    "field": "<step>",
                    "label": "步骤",
                    "issue": "missing_in_baseline",
                    "baseline": None,
                    "candidate": candidate_step.get("name"),
                }
            )
            continue
        action_name = _safe_text(_get_nested_dict_value(candidate_step, "actionConfig.action"))
        for spec in ROUNDTRIP_AUDIT_FIELD_SPECS:
            supported_actions = spec.get("actions")
            if supported_actions and action_name not in supported_actions:
                continue
            required_when = spec.get("required_when")
            if callable(required_when) and not required_when(candidate_step):
                continue
            field_path = spec["field"]
            baseline_value = _get_nested_dict_value(baseline_step, field_path)
            candidate_value = _get_nested_dict_value(candidate_step, field_path)
            if _normalize_audit_value(baseline_value) == _normalize_audit_value(candidate_value):
                continue
            issues.append(
                {
                    "stepId": step_id,
                    "stepName": _safe_text(candidate_step.get("name") or baseline_step.get("name")),
                    "field": field_path,
                    "label": spec.get("label") or field_path,
                    "issue": "value_changed",
                    "baseline": baseline_value,
                    "candidate": candidate_value,
                }
            )
    return issues


def format_roundtrip_audit_report(
    issues,
    baseline_label="基线",
    candidate_label="导入结果",
    max_entries=40,
):
    issues = list(issues or [])
    if not issues:
        return "关键回灌字段检查通过，未发现差异。"
    lines = [
        "检测到关键回灌字段差异：{count} 项".format(count=len(issues)),
        "对比基线：{baseline}".format(baseline=baseline_label or "未命名"),
        "导入结果：{candidate}".format(candidate=candidate_label or "未命名"),
        "",
    ]
    for item in issues[: max_entries or len(issues)]:
        lines.extend(
            [
                "步骤 {step_id} | {step_name}".format(
                    step_id=item.get("stepId", ""),
                    step_name=item.get("stepName", "") or "未命名步骤",
                ),
                "字段：{label}".format(label=item.get("label", item.get("field", ""))),
                "基线：{baseline}".format(
                    baseline=_normalize_audit_value(item.get("baseline")) or "<空>"
                ),
                "导入：{candidate}".format(
                    candidate=_normalize_audit_value(item.get("candidate")) or "<空>"
                ),
                "",
            ]
        )
    remaining = len(issues) - min(len(issues), max_entries or len(issues))
    if remaining > 0:
        lines.append("其余 {count} 项差异已省略，请改用命令行 audit 查看完整结果。".format(count=remaining))
    return "\n".join(lines).strip()


def audit_flow_excel_roundtrip(
    excel_path=DEFAULT_FLOW_XLSX,
    candidate_payload=None,
    baseline_path=None,
    step_ids=None,
    candidate_label="Excel 导入结果",
):
    meta = load_flow_excel_meta(excel_path)
    resolved_baseline_path = _safe_text(baseline_path) or _safe_text(meta.get(ROUNDTRIP_AUDIT_BASELINE_PATH_FIELD))
    resolved_step_ids = list(step_ids or []) or _parse_roundtrip_audit_step_ids(
        meta.get(ROUNDTRIP_AUDIT_STEP_IDS_FIELD)
    )
    baseline_label = (
        _safe_text(meta.get(ROUNDTRIP_AUDIT_BASELINE_LABEL_FIELD))
        or os.path.basename(resolved_baseline_path)
        or "导出基线"
    )
    result = {
        "available": False,
        "hasIssues": False,
        "issues": [],
        "report": "",
        "reason": "",
        "baselinePath": resolved_baseline_path,
        "baselineLabel": baseline_label,
        "candidateLabel": _safe_text(candidate_label) or "Excel 导入结果",
        "stepIds": resolved_step_ids,
    }
    if not resolved_baseline_path:
        result["reason"] = "Excel meta 页未记录基线链路路径，无法自动执行回灌审计。"
        result["report"] = result["reason"]
        return result
    baseline_abspath = os.path.abspath(resolved_baseline_path)
    if not os.path.exists(baseline_abspath):
        result["reason"] = f"未找到回灌审计基线文件：{baseline_abspath}"
        result["report"] = result["reason"]
        result["baselinePath"] = baseline_abspath
        return result
    baseline_payload = _read_json_file(baseline_abspath)
    if candidate_payload is None:
        candidate_payload = load_flow_payload_from_excel(excel_path)
    issues = compare_roundtrip_fields_between_payloads(
        baseline_payload,
        candidate_payload,
        step_ids=resolved_step_ids or None,
    )
    result["available"] = True
    result["hasIssues"] = bool(issues)
    result["issues"] = issues
    result["baselinePath"] = baseline_abspath
    result["report"] = format_roundtrip_audit_report(
        issues,
        baseline_label=baseline_label,
        candidate_label=result["candidateLabel"],
    )
    return result


def _json_dumps(value):
    if value in (None, "", [], {}):
        return ""
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def _json_loads(text, default_value):
    raw_text = str(text or "").strip()
    if not raw_text:
        return default_value
    try:
        return json.loads(raw_text)
    except json.JSONDecodeError as exc:
        raise ValueError(f"JSON 字段解析失败：{raw_text}") from exc


def _to_bool(value):
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    return text in {"1", "true", "yes", "y", "是"}


def _safe_text(value):
    return "" if value is None else str(value).strip()


def _safe_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        numeric = float(text)
    except Exception:
        return text
    if numeric.is_integer():
        return int(numeric)
    return numeric


def _sort_order_value(value):
    numeric_value = _safe_number(value)
    if isinstance(numeric_value, (int, float)):
        return (0, numeric_value)
    text_value = _safe_text(value)
    if text_value:
        return (1, text_value)
    return (2, "")


def _sheet_to_dicts(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(item or "").strip() for item in rows[0]]
    data_rows = []
    for row in rows[1:]:
        if not any(item not in (None, "") for item in row):
            continue
        item = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            item[header] = row[index] if index < len(row) else None
        data_rows.append(item)
    return data_rows


def _set_sheet_headers(sheet, headers):
    sheet.append(headers)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"


def _classify_header_fill(header_name):
    if header_name in CORE_REQUIRED_STEP_COLUMNS:
        return HEADER_FILL_CORE_REQUIRED
    if header_name in HIGH_FREQUENCY_ENGINEERING_COLUMNS:
        return HEADER_FILL_HIGH_FREQUENCY
    if header_name.endswith("Json"):
        return HEADER_FILL_JSON_FALLBACK
    return HEADER_FILL_OPTIONAL


def _annotate_sheet_headers(sheet, comment_map=None):
    comment_map = comment_map or {}
    for cell in sheet[1]:
        header_name = _safe_text(cell.value)
        if not header_name:
            continue
        cell.fill = PatternFill("solid", fgColor=_classify_header_fill(header_name))
        if header_name in comment_map:
            cell.comment = Comment(comment_map[header_name], "WT_Automation")


def _append_legend_rows(sheet, rows):
    if not rows:
        return
    start_row = sheet.max_row + 2
    for index, row in enumerate(rows, start=start_row):
        sheet.cell(row=index, column=1, value=row[0])
        sheet.cell(row=index, column=2, value=row[1])
        sheet.cell(row=index, column=1).font = Font(bold=True)
        sheet.cell(row=index, column=1).fill = PatternFill("solid", fgColor=row[2])
        sheet.cell(row=index, column=1).alignment = Alignment(vertical="center")
        sheet.cell(row=index, column=2).alignment = Alignment(vertical="top", wrap_text=True)


def _autofit_sheet(sheet, width_map=None):
    width_map = width_map or {}
    for column_cells in sheet.columns:
        letter = column_cells[0].column_letter
        header = str(column_cells[0].value or "")
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[letter].width = width_map.get(header, min(max(max_length + 2, 12), 48))
    if sheet.max_column and sheet.max_row:
        sheet.auto_filter.ref = sheet.dimensions


def _normalize_payload(payload):
    payload = payload if isinstance(payload, dict) else {}
    runtime_config = payload.get("runtimeConfig", {})
    flow_packages = payload.get("flowPackages", [])
    steps = payload.get("steps", [])
    if not isinstance(runtime_config, dict):
        runtime_config = {}
    if not isinstance(flow_packages, list):
        flow_packages = []
    if not isinstance(steps, list):
        steps = []
    normalized_steps = []
    for index, step in enumerate(steps, start=1):
        if not isinstance(step, dict):
            continue
        normalized_steps.append(
            {
                "seq": index,
                "id": _safe_text(step.get("id")),
                "name": _safe_text(step.get("name")),
                "stage": _safe_text(step.get("stage")),
                "strategy": _safe_text(step.get("strategy")),
                "actionType": _safe_text(step.get("actionType")) or "script",
                "topLevel": bool(step.get("topLevel", True)),
                "enabled": bool(step.get("enabled", True)),
                "packageRef": _safe_text(step.get("packageRef")),
                "codeSymbol": _safe_text(step.get("codeSymbol")),
                "codeReference": _safe_text(step.get("codeReference")),
                "windowTitle": _safe_text(step.get("windowTitle")),
                "successLog": _safe_text(step.get("successLog")),
                "description": _safe_text(step.get("description")),
                "notes": _safe_text(step.get("notes")),
                "inspectHints": step.get("inspectHints", {}) if isinstance(step.get("inspectHints"), dict) else {},
                "stepParams": step.get("stepParams", {}) if isinstance(step.get("stepParams"), dict) else {},
                "actionConfig": step.get("actionConfig", {}) if isinstance(step.get("actionConfig"), dict) else {},
                "auxChecks": step.get("auxChecks", []) if isinstance(step.get("auxChecks"), list) else [],
                "fallbacks": step.get("fallbacks", []) if isinstance(step.get("fallbacks"), list) else [],
                "fallbackChain": step.get("fallbackChain", []) if isinstance(step.get("fallbackChain"), list) else [],
                "controls": step.get("controls", []) if isinstance(step.get("controls"), list) else [],
            }
        )
    return {
        "version": _safe_text(payload.get("version")) or "1.0",
        "project": _safe_text(payload.get("project")) or "WT_Automation",
        "description": _safe_text(payload.get("description")),
        "lastUpdated": _safe_text(payload.get("lastUpdated")),
        "runtimeConfig": {
            # 保留全部运行时配置键（含 controlMapPath / templateAutoUpdate 等扩展字段），
            # 仅做基础类型清洗（bool/数字保留原类型），
            # 避免白名单重建导致执行器依赖的配置在 Excel 往返中静默丢失或类型漂移。
            key: (
                value
                if isinstance(value, (bool, int, float, dict, list))
                else _safe_text(value)
            )
            for key, value in runtime_config.items()
        },
        "flowPackages": [
            {
                "id": _safe_text(package.get("id")),
                "name": _safe_text(package.get("name")),
                "description": _safe_text(package.get("description")),
                "stepIds": [str(item).strip() for item in package.get("stepIds", []) if str(item).strip()],
            }
            for package in flow_packages
            if isinstance(package, dict) and _safe_text(package.get("id"))
        ],
        "steps": normalized_steps,
    }


def _filter_payload_by_step_ids(payload, selected_step_ids=None):
    normalized = _normalize_payload(payload)
    if not selected_step_ids:
        return normalized

    ordered_step_ids = []
    seen_step_ids = set()
    for step_id in selected_step_ids:
        normalized_step_id = _safe_text(step_id)
        if normalized_step_id and normalized_step_id not in seen_step_ids:
            ordered_step_ids.append(normalized_step_id)
            seen_step_ids.add(normalized_step_id)
    if not ordered_step_ids:
        return normalized

    selected_set = set(ordered_step_ids)
    step_map = {
        _safe_text(step.get("id")): step
        for step in normalized.get("steps", [])
        if isinstance(step, dict) and _safe_text(step.get("id"))
    }
    filtered_steps = []
    for index, step_id in enumerate(ordered_step_ids, start=1):
        step = step_map.get(step_id)
        if not isinstance(step, dict):
            continue
        step_copy = dict(step)
        step_copy["seq"] = index
        filtered_steps.append(step_copy)

    filtered_packages = []
    filtered_step_ids = {step.get("id", "") for step in filtered_steps}
    for package in normalized.get("flowPackages", []):
        if not isinstance(package, dict):
            continue
        package_id = _safe_text(package.get("id"))
        if not package_id:
            continue
        package_step_ids = [
            _safe_text(item)
            for item in (package.get("stepIds") or [])
            if _safe_text(item) in filtered_step_ids
        ]
        has_step_ref = any(_safe_text(step.get("packageRef")) == package_id for step in filtered_steps)
        if not package_step_ids and not has_step_ref:
            continue
        package_copy = dict(package)
        package_copy["stepIds"] = package_step_ids
        filtered_packages.append(package_copy)

    normalized["steps"] = filtered_steps
    normalized["flowPackages"] = filtered_packages
    return normalized


def _extract_engineering_step_fields(step):
    inspect_hints = step.get("inspectHints", {}) if isinstance(step.get("inspectHints"), dict) else {}
    action_config = step.get("actionConfig", {}) if isinstance(step.get("actionConfig"), dict) else {}
    parent_window = action_config.get("parentWindow", {}) if isinstance(action_config.get("parentWindow"), dict) else {}
    relative_region = action_config.get("relativeRegion", {}) if isinstance(action_config.get("relativeRegion"), dict) else {}
    continue_when = action_config.get("continueWhen", {}) if isinstance(action_config.get("continueWhen"), dict) else {}
    action_config_extra = {
        key: value
        for key, value in action_config.items()
        if key not in KNOWN_ACTION_CONFIG_KEYS
    }
    input_text = ""
    if "text" in action_config:
        input_text = _safe_text(action_config.get("text"))
    elif "value" in action_config:
        input_text = _safe_text(action_config.get("value"))
    return {
        "inspectControlName": _safe_text(inspect_hints.get("controlName")),
        "inspectClassName": _safe_text(inspect_hints.get("className")),
        "inspectAutomationId": _safe_text(inspect_hints.get("automationId")),
        "inspectControlType": _safe_text(inspect_hints.get("controlType")),
        "inspectUiPath": _safe_text(inspect_hints.get("uiPath")),
        "inspectTemplateKey": _safe_text(inspect_hints.get("templateKey")),
        "action": _safe_text(action_config.get("action")),
        "controlId": _safe_text(action_config.get("controlId")),
        "inputText": input_text,
        "postInputKeys": _safe_text(action_config.get("postInputKeys")),
        "waitSeconds": action_config.get("seconds", ""),
        "wheelDelta": action_config.get("delta", ""),
        "timeoutSeconds": action_config.get("timeoutSeconds", ""),
        "waitBefore": action_config.get("waitBefore", ""),
        "waitAfter": action_config.get("waitAfter", ""),
        "continueWhenControlId": _safe_text(continue_when.get("controlId")),
        "continueWhenCondition": _safe_text(continue_when.get("condition")),
        "continueWhenTimeoutSeconds": continue_when.get("timeoutSeconds", ""),
        "continueWhenWindowTitleHint": _safe_text(continue_when.get("windowTitleHint")),
        "retryCount": action_config.get("retryCount", ""),
        "retryInterval": action_config.get("retryInterval", ""),
        "onError": _safe_text(action_config.get("onError")),
        "fallbackMode": _safe_text(action_config.get("fallbackMode")),
        "fallbackTemplate": _safe_text(action_config.get("fallbackTemplate")),
        "saveAs": _safe_text(action_config.get("saveAs")),
        "parentWindowTitle": _safe_text(parent_window.get("title")),
        "parentWindowClassName": _safe_text(parent_window.get("className")),
        "parentWindowFrameworkId": _safe_text(parent_window.get("frameworkId")),
        "regionX": relative_region.get("x", ""),
        "regionY": relative_region.get("y", ""),
        "regionWidth": relative_region.get("width", ""),
        "regionHeight": relative_region.get("height", ""),
        "regionAnchor": _safe_text(relative_region.get("anchor")),
        "inspectHintsJson": _json_dumps(
            {
                key: value
                for key, value in inspect_hints.items()
                if key
                not in {"controlName", "className", "automationId", "controlType", "uiPath", "templateKey"}
            }
        ),
        "actionConfigJson": _json_dumps(action_config_extra),
    }


def _build_option_lists(payload):
    observed_strategies = []
    observed_target_methods = []
    for step in payload.get("steps", []):
        strategy = _safe_text(step.get("strategy"))
        if strategy and strategy not in observed_strategies:
            observed_strategies.append(strategy)
        for control in step.get("controls", []):
            if not isinstance(control, dict):
                continue
            target_method = _safe_text(control.get("targetMethod"))
            if target_method and target_method not in observed_target_methods:
                observed_target_methods.append(target_method)
    strategy_options = [item for item in STRATEGY_OPTIONS if item]
    for item in observed_strategies:
        if item not in strategy_options:
            strategy_options.append(item)
    target_method_options = list(COMMON_TARGET_METHOD_OPTIONS)
    for item in observed_target_methods:
        if item not in target_method_options:
            target_method_options.append(item)
    package_options = [package.get("id", "") for package in payload.get("flowPackages", []) if package.get("id", "")]
    return {
        "strategy": strategy_options,
        "actionType": ACTION_TYPE_OPTIONS,
        "boolean": BOOLEAN_OPTIONS,
        "action": list(get_action_names()),
        "onError": ON_ERROR_OPTIONS,
        "fallbackMode": FALLBACK_MODE_OPTIONS,
        "frameworkId": FRAMEWORK_OPTIONS,
        "anchor": ANCHOR_OPTIONS,
        "continueCondition": CONTINUE_CONDITION_OPTIONS,
        "packageRef": package_options,
        "targetMethod": target_method_options,
    }


def _create_action_guide_sheet(workbook):
    guide_sheet = workbook.create_sheet(ACTION_GUIDE_SHEET_NAME)
    headers = ["action", "label", "必填目标", "必填输入", "建议填写列", "说明"]
    _set_sheet_headers(guide_sheet, headers)
    for action_name in get_action_names():
        schema = get_action_schema(action_name)
        suggested_columns = []
        if schema.get("target_required"):
            suggested_columns.append("controlId + controls.targetMethod/targetValue")
        if schema.get("input_required"):
            input_key = _safe_text(schema.get("input_key"))
            if input_key in {"text", "value"}:
                suggested_columns.append("inputText")
            elif input_key == "seconds":
                suggested_columns.append("waitSeconds")
            elif input_key == "delta":
                suggested_columns.append("wheelDelta")
        for column_name in schema.get("suggested_columns", ()) or ():
            if column_name and column_name not in suggested_columns:
                suggested_columns.append(column_name)
        if action_name in {"click_relative_region", "type_text_relative"}:
            composite_columns = "parentWindowTitle/className/frameworkId + regionX/Y/Width/Height"
            if composite_columns not in suggested_columns:
                suggested_columns.append(composite_columns)
        if action_name not in {"sleep", "log"}:
            continue_columns = "continueWhenControlId/Condition/TimeoutSeconds"
            if continue_columns not in suggested_columns:
                suggested_columns.append(continue_columns)
        guide_sheet.append(
            [
                action_name,
                _safe_text(schema.get("label")),
                "是" if schema.get("target_required") else "否",
                "是" if schema.get("input_required") else "否",
                "；".join(suggested_columns) or "无额外列",
                build_action_schema_hint(action_name),
            ]
        )
    _autofit_sheet(
        guide_sheet,
        {
            "action": 28,
            "label": 18,
            "建议填写列": 46,
            "说明": 68,
        },
    )
    return guide_sheet


def _create_examples_sheet(workbook):
    examples_sheet = workbook.create_sheet(EXAMPLES_SHEET_NAME)
    _set_sheet_headers(examples_sheet, STEP_COLUMNS)
    examples = [
        {
            "seq": 1,
            "id": "example_click_button",
            "name": "点击-添加到数据",
            "strategy": "action",
            "actionType": "action",
            "enabled": "是",
            "action": "click",
            "controlId": "control_add_data",
            "continueWhenControlId": "control_data_added_flag",
            "continueWhenCondition": "visible",
            "continueWhenTimeoutSeconds": 8,
            "description": "示例：标准控件点击，目标控件在 controls 表维护；点击后等待结果控件出现再续跑。",
        },
        {
            "seq": 2,
            "id": "example_type_text_relative",
            "name": "输入-默认高度",
            "strategy": "action",
            "actionType": "action",
            "enabled": "是",
            "action": "type_text_relative",
            "inputText": "100",
            "postInputKeys": "{TAB}",
            "parentWindowTitle": "",
            "parentWindowClassName": "Window",
            "parentWindowFrameworkId": "WPF",
            "regionX": 0.43,
            "regionY": 0.56,
            "regionWidth": 0.08,
            "regionHeight": 0.03,
            "regionAnchor": "center",
            "continueWhenControlId": "control_height_confirm",
            "continueWhenCondition": "enabled",
            "continueWhenTimeoutSeconds": 6,
            "description": "示例：父窗口相对区域输入，适合 WPF 主界面；输入后等待确认按钮可用。",
        },
        {
            "seq": 3,
            "id": "example_select_dropdown_runtime",
            "name": "下拉-选择日期时间",
            "strategy": "action",
            "actionType": "action",
            "enabled": "是",
            "action": "select_dropdown_item_runtime",
            "controlId": "control_datetime_item",
            "continueWhenControlId": "control_datetime_selected_tag",
            "continueWhenCondition": "visible",
            "continueWhenTimeoutSeconds": 5,
            "description": "示例：WPF 下拉项运行时选择，先点击展开步骤，再选项步骤，并等待选中结果出现。",
        },
    ]
    for example in examples:
        examples_sheet.append([example.get(column, "") for column in STEP_COLUMNS])
    _annotate_sheet_headers(examples_sheet, STEP_COLUMN_COMMENTS)
    _autofit_sheet(
        examples_sheet,
        {
            "id": 28,
            "name": 24,
            "description": 58,
            "parentWindowTitle": 24,
            "parentWindowClassName": 24,
            "inspectHintsJson": 36,
            "actionConfigJson": 36,
        },
    )
    return examples_sheet


def _write_options_sheet(workbook, option_lists):
    options_sheet = workbook.create_sheet(OPTIONS_SHEET_NAME)
    option_ranges = {}
    for column_index, (key, values) in enumerate(option_lists.items(), start=1):
        options_sheet.cell(row=1, column=column_index, value=key)
        for row_index, value in enumerate(values, start=2):
            options_sheet.cell(row=row_index, column=column_index, value=value)
        column_letter = get_column_letter(column_index)
        end_row = max(2, len(values) + 1)
        option_ranges[key] = f"='{OPTIONS_SHEET_NAME}'!${column_letter}$2:${column_letter}${end_row}"
    options_sheet.sheet_state = "hidden"
    return option_ranges


def _header_to_column_letter(sheet):
    mapping = {}
    for index, cell in enumerate(sheet[1], start=1):
        header = _safe_text(cell.value)
        if header:
            mapping[header] = get_column_letter(index)
    return mapping


def _apply_list_validation(sheet, header_map, header_name, formula, max_rows=2000):
    if not formula or header_name not in header_map:
        return
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"{header_map[header_name]}2:{header_map[header_name]}{max_rows}")


def export_flow_to_excel(flow_json_path=DEFAULT_FLOW_JSON, excel_path=DEFAULT_FLOW_XLSX, selected_step_ids=None):
    _require_openpyxl()
    payload = _filter_payload_by_step_ids(_read_json_file(flow_json_path), selected_step_ids=selected_step_ids)
    exported_step_ids = [
        _safe_text(step.get("id"))
        for step in payload.get("steps", [])
        if isinstance(step, dict) and _safe_text(step.get("id"))
    ]

    workbook = Workbook()
    meta_sheet = workbook.active
    meta_sheet.title = "meta"
    _set_sheet_headers(meta_sheet, ["field", "value"])
    meta_rows = [
        ("version", payload.get("version", "")),
        ("project", payload.get("project", "")),
        ("description", payload.get("description", "")),
        ("lastUpdated", payload.get("lastUpdated", "")),
        ("generatedAt", datetime.now().isoformat(timespec="seconds")),
    ]
    # runtimeConfig 全部键写入 meta（runtime.<key>），导入时按相同前缀读回，
    # 保证 controlMapPath / templateAutoUpdate 等扩展字段在 Excel 往返中不丢失。
    for key, value in payload.get("runtimeConfig", {}).items():
        if not isinstance(value, (dict, list)):
            meta_rows.append(("runtime." + str(key), value))
    meta_rows.extend([
        (ROUNDTRIP_AUDIT_BASELINE_PATH_FIELD, os.path.abspath(flow_json_path)),
        (ROUNDTRIP_AUDIT_BASELINE_LABEL_FIELD, os.path.basename(flow_json_path)),
        (ROUNDTRIP_AUDIT_STEP_IDS_FIELD, json.dumps(exported_step_ids, ensure_ascii=False)),
    ])
    for row in meta_rows:
        meta_sheet.append(list(row))
    _annotate_sheet_headers(meta_sheet, {"field": "固定字段名，请不要修改。", "value": "项目级配置值，可直接编辑。"})

    steps_sheet = workbook.create_sheet("steps")
    _set_sheet_headers(steps_sheet, STEP_COLUMNS)
    for step in payload["steps"]:
        engineering = _extract_engineering_step_fields(step)
        steps_sheet.append(
            [
                step["seq"],
                step["id"],
                step["name"],
                step["stage"],
                step["strategy"],
                step["actionType"],
                "是" if step["topLevel"] else "否",
                "是" if step["enabled"] else "否",
                step["packageRef"],
                step["codeSymbol"],
                step["codeReference"],
                step["windowTitle"],
                step["successLog"],
                step["description"],
                step["notes"],
                engineering["inspectControlName"],
                engineering["inspectClassName"],
                engineering["inspectAutomationId"],
                engineering["inspectControlType"],
                engineering["inspectUiPath"],
                engineering["inspectTemplateKey"],
                engineering["action"],
                engineering["controlId"],
                engineering["inputText"],
                engineering["postInputKeys"],
                engineering["waitSeconds"],
                engineering["wheelDelta"],
                engineering["timeoutSeconds"],
                engineering["waitBefore"],
                engineering["waitAfter"],
                engineering["continueWhenControlId"],
                engineering["continueWhenCondition"],
                engineering["continueWhenTimeoutSeconds"],
                engineering["continueWhenWindowTitleHint"],
                engineering["retryCount"],
                engineering["retryInterval"],
                engineering["onError"],
                engineering["fallbackMode"],
                engineering["fallbackTemplate"],
                engineering["saveAs"],
                engineering["parentWindowTitle"],
                engineering["parentWindowClassName"],
                engineering["parentWindowFrameworkId"],
                engineering["regionX"],
                engineering["regionY"],
                engineering["regionWidth"],
                engineering["regionHeight"],
                engineering["regionAnchor"],
                engineering["inspectHintsJson"],
                _json_dumps(step["stepParams"]),
                engineering["actionConfigJson"],
                _json_dumps(step["auxChecks"]),
                _json_dumps(step["fallbacks"]),
                _json_dumps(step.get("fallbackChain", [])),
            ]
        )
    _annotate_sheet_headers(steps_sheet, STEP_COLUMN_COMMENTS)

    controls_sheet = workbook.create_sheet("controls")
    _set_sheet_headers(controls_sheet, CONTROL_COLUMNS)
    for step in payload["steps"]:
        for index, control in enumerate(step.get("controls", []), start=1):
            if not isinstance(control, dict):
                continue
            controls_sheet.append(
                [
                    step["id"],
                    index,
                    _safe_text(control.get("id")),
                    _safe_text(control.get("name")),
                    _safe_text(control.get("role")),
                    _safe_text(control.get("targetMethod")),
                    _safe_text(control.get("targetValue")),
                    _safe_text(control.get("windowTitle")),
                    _safe_text(control.get("templateKey")),
                    _safe_text(control.get("targetIndex")),
                    _json_dumps(control.get("inspectData", {}) if isinstance(control.get("inspectData"), dict) else {}),
                    _json_dumps(control.get("auxChecks", []) if isinstance(control.get("auxChecks"), list) else []),
                    _safe_text(control.get("notes")),
                ]
            )
    _annotate_sheet_headers(controls_sheet, CONTROL_COLUMN_COMMENTS)

    packages_sheet = workbook.create_sheet("flow_packages")
    _set_sheet_headers(packages_sheet, PACKAGE_COLUMNS)
    for package in payload["flowPackages"]:
        packages_sheet.append(
            [
                package.get("id", ""),
                package.get("name", ""),
                package.get("description", ""),
                _json_dumps(package.get("stepIds", [])),
            ]
        )
    _annotate_sheet_headers(
        packages_sheet,
        {
            "id": "流程包唯一标识。",
            "name": "流程包名称，供主界面展示。",
            "stepIdsJson": "步骤 ID 数组，例如 [\"step_1\",\"step_2\"]。",
        },
    )

    guide_sheet = workbook.create_sheet("guide")
    _set_sheet_headers(guide_sheet, ["sheet", "说明"])
    guide_rows = [
        ("meta", "项目级信息与运行参数。修改 value 列即可。"),
        ("steps", "一行一个步骤。优先编辑显式列；inspectHintsJson / actionConfigJson 仅用于高级补充配置。"),
        ("controls", "一行一个控件，按 stepId 关联步骤；targetMethod 支持下拉常用定位组合。"),
        ("flow_packages", "流程包定义，stepIdsJson 填 JSON 数组。"),
        (ACTION_GUIDE_SHEET_NAME, "动作规则总表。按 action 查看该动作需要目标控件还是输入参数，以及建议填写哪些列。"),
        (EXAMPLES_SHEET_NAME, "工程师示例表。展示常见动作的标准填写方式，可直接照着抄。"),
        (OPTIONS_SHEET_NAME, "隐藏选项表，为 steps / controls 提供下拉选项，请勿手动删除。"),
    ]
    for row in guide_rows:
        guide_sheet.append(list(row))
    _append_legend_rows(
        guide_sheet,
        [
            ("上手顺序", "建议先看 examples，再看 action_guide，最后回到 steps 和 controls 维护正式流程。", HEADER_FILL_HIGH_FREQUENCY),
            ("颜色说明", "黄色=核心必填；蓝色=工程师高频维护列；紫色=高级 JSON 兜底列。", HEADER_FILL_CORE_REQUIRED),
            ("填写建议", "优先维护显式列；只有显式列无法覆盖时，才填写 inspectHintsJson / actionConfigJson。", HEADER_FILL_JSON_FALLBACK),
        ],
    )

    _create_action_guide_sheet(workbook)
    _create_examples_sheet(workbook)

    option_ranges = _write_options_sheet(workbook, _build_option_lists(payload))
    steps_header_map = _header_to_column_letter(steps_sheet)
    controls_header_map = _header_to_column_letter(controls_sheet)
    _apply_list_validation(steps_sheet, steps_header_map, "strategy", option_ranges.get("strategy"))
    _apply_list_validation(steps_sheet, steps_header_map, "actionType", option_ranges.get("actionType"))
    _apply_list_validation(steps_sheet, steps_header_map, "topLevel", option_ranges.get("boolean"))
    _apply_list_validation(steps_sheet, steps_header_map, "enabled", option_ranges.get("boolean"))
    _apply_list_validation(steps_sheet, steps_header_map, "action", option_ranges.get("action"))
    _apply_list_validation(steps_sheet, steps_header_map, "onError", option_ranges.get("onError"))
    _apply_list_validation(steps_sheet, steps_header_map, "fallbackMode", option_ranges.get("fallbackMode"))
    _apply_list_validation(steps_sheet, steps_header_map, "parentWindowFrameworkId", option_ranges.get("frameworkId"))
    _apply_list_validation(steps_sheet, steps_header_map, "regionAnchor", option_ranges.get("anchor"))
    _apply_list_validation(steps_sheet, steps_header_map, "continueWhenCondition", option_ranges.get("continueCondition"))
    _apply_list_validation(steps_sheet, steps_header_map, "packageRef", option_ranges.get("packageRef"))
    _apply_list_validation(controls_sheet, controls_header_map, "targetMethod", option_ranges.get("targetMethod"))

    _autofit_sheet(meta_sheet, {"field": 28, "value": 72})
    _autofit_sheet(
        steps_sheet,
        {
            "id": 20,
            "name": 22,
            "description": 34,
            "notes": 32,
            "windowTitle": 26,
            "successLog": 30,
            "inputText": 24,
            "postInputKeys": 16,
            "fallbackTemplate": 30,
            "continueWhenWindowTitleHint": 28,
            "inspectUiPath": 34,
            "inspectHintsJson": 42,
            "stepParamsJson": 34,
            "actionConfigJson": 42,
            "auxChecksJson": 28,
            "fallbacksJson": 24,
        },
    )
    _autofit_sheet(
        controls_sheet,
        {
            "stepId": 18,
            "id": 24,
            "name": 22,
            "role": 26,
            "targetMethod": 22,
            "targetValue": 40,
            "inspectDataJson": 48,
            "auxChecksJson": 28,
            "notes": 28,
        },
    )
    _autofit_sheet(packages_sheet, {"description": 40, "stepIdsJson": 42})
    _autofit_sheet(guide_sheet, {"sheet": 18, "说明": 80})

    workbook.save(excel_path)
    return excel_path


def _merge_step_inspect_hints(row):
    inspect_hints = _json_loads(row.get("inspectHintsJson"), {})
    if not isinstance(inspect_hints, dict):
        inspect_hints = {}
    explicit_map = {
        "controlName": _safe_text(row.get("inspectControlName")),
        "className": _safe_text(row.get("inspectClassName")),
        "automationId": _safe_text(row.get("inspectAutomationId")),
        "controlType": _safe_text(row.get("inspectControlType")),
        "uiPath": _safe_text(row.get("inspectUiPath")),
        "templateKey": _safe_text(row.get("inspectTemplateKey")),
    }
    for key, value in explicit_map.items():
        if value:
            inspect_hints[key] = value
        elif key in inspect_hints and not value and key in {
            "controlName",
            "className",
            "automationId",
            "controlType",
            "uiPath",
            "templateKey",
        }:
            inspect_hints.pop(key, None)
    return inspect_hints


def _merge_step_action_config(row):
    action_config = _json_loads(row.get("actionConfigJson"), {})
    if not isinstance(action_config, dict):
        action_config = {}
    action_name = _safe_text(row.get("action")) or _safe_text(action_config.get("action"))
    if action_name:
        action_config["action"] = action_name
    control_id = _safe_text(row.get("controlId"))
    if control_id:
        action_config["controlId"] = control_id
    input_text = _safe_text(row.get("inputText"))
    if input_text:
        action_config["text"] = input_text
    post_input_keys = _safe_text(row.get("postInputKeys"))
    if post_input_keys:
        action_config["postInputKeys"] = post_input_keys
    else:
        action_config.pop("postInputKeys", None)
    seconds_value = _safe_number(row.get("waitSeconds"))
    if seconds_value is not None:
        action_config["seconds"] = seconds_value
    delta_value = _safe_number(row.get("wheelDelta"))
    if delta_value is not None:
        action_config["delta"] = delta_value
    for source_key, target_key in [
        ("timeoutSeconds", "timeoutSeconds"),
        ("waitBefore", "waitBefore"),
        ("waitAfter", "waitAfter"),
        ("retryCount", "retryCount"),
        ("retryInterval", "retryInterval"),
    ]:
        numeric_value = _safe_number(row.get(source_key))
        if numeric_value is not None:
            action_config[target_key] = numeric_value
    for text_key in ("onError", "fallbackMode", "fallbackTemplate", "saveAs"):
        text_value = _safe_text(row.get(text_key))
        if text_value:
            action_config[text_key] = text_value
    parent_window = action_config.get("parentWindow", {}) if isinstance(action_config.get("parentWindow"), dict) else {}
    parent_title = _safe_text(row.get("parentWindowTitle"))
    parent_class_name = _safe_text(row.get("parentWindowClassName"))
    parent_framework = _safe_text(row.get("parentWindowFrameworkId"))
    if parent_title:
        parent_window["title"] = parent_title
    if parent_class_name:
        parent_window["className"] = parent_class_name
    if parent_framework:
        parent_window["frameworkId"] = parent_framework
    if parent_window:
        action_config["parentWindow"] = parent_window
    relative_region = action_config.get("relativeRegion", {}) if isinstance(action_config.get("relativeRegion"), dict) else {}
    for source_key, target_key in [
        ("regionX", "x"),
        ("regionY", "y"),
        ("regionWidth", "width"),
        ("regionHeight", "height"),
    ]:
        numeric_value = _safe_number(row.get(source_key))
        if numeric_value is not None:
            relative_region[target_key] = numeric_value
    anchor = _safe_text(row.get("regionAnchor"))
    if anchor:
        relative_region["anchor"] = anchor
    if relative_region:
        action_config["relativeRegion"] = relative_region
    continue_when = action_config.get("continueWhen", {}) if isinstance(action_config.get("continueWhen"), dict) else {}
    continue_control_id = _safe_text(row.get("continueWhenControlId"))
    continue_condition = _safe_text(row.get("continueWhenCondition"))
    continue_window_title_hint = _safe_text(row.get("continueWhenWindowTitleHint"))
    continue_timeout = _safe_number(row.get("continueWhenTimeoutSeconds"))
    if continue_control_id:
        continue_when["controlId"] = continue_control_id
    if continue_condition:
        continue_when["condition"] = continue_condition
    if continue_window_title_hint:
        continue_when["windowTitleHint"] = continue_window_title_hint
    if continue_timeout is not None:
        continue_when["timeoutSeconds"] = continue_timeout
    if continue_when:
        action_config["continueWhen"] = continue_when
    return action_config


def _cleanup_step_payload(step):
    for key in ["inspectHints", "stepParams", "actionConfig"]:
        if step.get(key) == {}:
            step.pop(key, None)
    for key in ["auxChecks", "fallbacks", "fallbackChain", "controls"]:
        if step.get(key) == []:
            step.pop(key, None)
    for key in ["notes", "packageRef", "codeSymbol", "codeReference", "description", "successLog", "windowTitle", "stage", "strategy"]:
        if key in step and not _safe_text(step.get(key)):
            step.pop(key, None)
    return step


def load_flow_payload_from_excel(excel_path=DEFAULT_FLOW_XLSX):
    _require_openpyxl()
    workbook = load_workbook(excel_path)
    try:
        if "meta" not in workbook.sheetnames or "steps" not in workbook.sheetnames:
            raise ValueError("Excel 缺少必要工作表：meta 或 steps。")

        meta_rows = _sheet_to_dicts(workbook["meta"])
        meta_map = {_safe_text(row.get("field")): row.get("value") for row in meta_rows if _safe_text(row.get("field"))}

        controls_by_step = {}
        if "controls" in workbook.sheetnames:
            for row in _sheet_to_dicts(workbook["controls"]):
                step_id = _safe_text(row.get("stepId"))
                control_id = _safe_text(row.get("id"))
                if not step_id or not control_id:
                    continue
                control = {
                    "id": control_id,
                    "name": _safe_text(row.get("name")),
                    "role": _safe_text(row.get("role")),
                    "targetMethod": _safe_text(row.get("targetMethod")),
                    "targetValue": _safe_text(row.get("targetValue")),
                }
                window_title = _safe_text(row.get("windowTitle"))
                template_key = _safe_text(row.get("templateKey"))
                target_index = _safe_text(row.get("targetIndex"))
                notes = _safe_text(row.get("notes"))
                inspect_data = _json_loads(row.get("inspectDataJson"), {})
                aux_checks = _json_loads(row.get("auxChecksJson"), [])
                if window_title:
                    control["windowTitle"] = window_title
                if template_key:
                    control["templateKey"] = template_key
                if target_index:
                    control["targetIndex"] = target_index
                if isinstance(inspect_data, dict) and inspect_data:
                    control["inspectData"] = inspect_data
                if isinstance(aux_checks, list) and aux_checks:
                    control["auxChecks"] = aux_checks
                if notes:
                    control["notes"] = notes
                controls_by_step.setdefault(step_id, []).append((row.get("order") or 0, control))

        steps = []
        for row in _sheet_to_dicts(workbook["steps"]):
            step_id = _safe_text(row.get("id"))
            if not step_id:
                continue
            step = {
                "id": step_id,
                "name": _safe_text(row.get("name")),
                "stage": _safe_text(row.get("stage")),
                "strategy": _safe_text(row.get("strategy")),
                "actionType": _safe_text(row.get("actionType")) or "script",
                "topLevel": _to_bool(row.get("topLevel")) if row.get("topLevel") not in (None, "") else True,
                "enabled": _to_bool(row.get("enabled")) if row.get("enabled") not in (None, "") else True,
                "codeSymbol": _safe_text(row.get("codeSymbol")),
                "codeReference": _safe_text(row.get("codeReference")),
                "description": _safe_text(row.get("description")),
                "successLog": _safe_text(row.get("successLog")),
                "windowTitle": _safe_text(row.get("windowTitle")),
                "inspectHints": _merge_step_inspect_hints(row),
                "stepParams": _json_loads(row.get("stepParamsJson"), {}),
                "actionConfig": _merge_step_action_config(row),
                "auxChecks": _json_loads(row.get("auxChecksJson"), []),
                "fallbacks": _json_loads(row.get("fallbacksJson"), []),
                "fallbackChain": _json_loads(row.get("fallbackChainJson"), []),
                "notes": _safe_text(row.get("notes")),
            }
            package_ref = _safe_text(row.get("packageRef"))
            if package_ref:
                step["packageRef"] = package_ref
            control_items = controls_by_step.get(step_id, [])
            if control_items:
                step["controls"] = [item[1] for item in sorted(control_items, key=lambda pair: (pair[0], pair[1].get("id", "")))]
            steps.append((row.get("seq") or 0, _cleanup_step_payload(step)))

        flow_packages = []
        if "flow_packages" in workbook.sheetnames:
            for row in _sheet_to_dicts(workbook["flow_packages"]):
                package_id = _safe_text(row.get("id"))
                if not package_id:
                    continue
                package = {
                    "id": package_id,
                    "name": _safe_text(row.get("name")) or package_id,
                    "description": _safe_text(row.get("description")),
                    "stepIds": _json_loads(row.get("stepIdsJson"), []),
                }
                flow_packages.append(package)

        runtime_config = {}
        for field_key, field_value in meta_map.items():
            field_key = _safe_text(field_key)
            if not field_key.startswith("runtime."):
                continue
            runtime_config[field_key[len("runtime."):]] = field_value

        payload = {
            "version": _safe_text(meta_map.get("version")) or "1.0",
            "project": _safe_text(meta_map.get("project")) or "WT_Automation",
            "description": _safe_text(meta_map.get("description")),
            "lastUpdated": datetime.now().isoformat(timespec="seconds"),
            "runtimeConfig": runtime_config,
            "flowPackages": flow_packages,
            "steps": [item[1] for item in sorted(steps, key=lambda pair: (_sort_order_value(pair[0]), pair[1].get("id", "")))],
        }
        return payload
    finally:
        try:
            workbook.close()
        except Exception:
            pass


def import_flow_from_excel(excel_path=DEFAULT_FLOW_XLSX, flow_json_path=DEFAULT_FLOW_JSON):
    payload = load_flow_payload_from_excel(excel_path)
    _write_json_file(flow_json_path, payload)
    return flow_json_path


def build_arg_parser():
    parser = argparse.ArgumentParser(description="WT 自动化流程 Excel 导入导出工具")
    subparsers = parser.add_subparsers(dest="command", required=True)

    export_parser = subparsers.add_parser("export", help="从 flow_definition.json 导出为 Excel")
    export_parser.add_argument("--json", default=DEFAULT_FLOW_JSON, help="输入 JSON 路径")
    export_parser.add_argument("--xlsx", default=DEFAULT_FLOW_XLSX, help="输出 Excel 路径")

    import_parser = subparsers.add_parser("import", help="从 Excel 导入为 flow_definition.json")
    import_parser.add_argument("--xlsx", default=DEFAULT_FLOW_XLSX, help="输入 Excel 路径")
    import_parser.add_argument("--json", default=DEFAULT_FLOW_JSON, help="输出 JSON 路径")

    return parser


def main():
    parser = build_arg_parser()
    args = parser.parse_args()
    if args.command == "export":
        result = export_flow_to_excel(args.json, args.xlsx)
        print(f"已导出流程 Excel: {result}")
        return
    if args.command == "import":
        result = import_flow_from_excel(args.xlsx, args.json)
        print(f"已导入流程定义: {result}")
        return
    parser.error("未知命令")


if __name__ == "__main__":
    main()
